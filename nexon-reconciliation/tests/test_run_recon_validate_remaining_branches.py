from __future__ import annotations

import builtins
import copy
import os
import subprocess
import sys
import tempfile
import unittest
import zipfile
from argparse import Namespace
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook


SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
TESTS = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS))
sys.path.insert(0, str(TESTS))

from recon_core import run_recon, validate_run  # noqa: E402
from recon_core.common import read_json, sha256_file, write_json  # noqa: E402
from test_orchestrator_validator_lifecycle import (  # noqa: E402
    CONFIG,
    RuntimeHarness,
    _all_capabilities,
    _args,
    _download_receipt,
    _sharepoint_binding,
)


def _new_run(
    root: Path,
    *,
    run_mode: str,
    matched: bool = False,
    local_only: bool = True,
    source_suffix: str = ".csv",
    core_mode: str = "sqlite_shadow",
    investigation: Path | None = None,
) -> tuple[dict, Path]:
    source = root / f"invoice{source_suffix}"
    if source_suffix == ".zip":
        with zipfile.ZipFile(source, "w") as archive:
            archive.writestr("invoice.csv", "invoice")
    else:
        source.write_text("invoice", encoding="utf-8")
    billing_sql = root / "billing.sql"
    billing_sql.write_text("SELECT 1", encoding="utf-8")
    result_root = root / "results"
    (result_root / "AAPT").mkdir(parents=True)
    harness = RuntimeHarness(matched)
    source_index = root / "sharepoint-file-index.json"
    source_index.write_text('{"contract_version":1}', encoding="utf-8")
    args = _args(
        source=source,
        result_root=result_root,
        run_mode=run_mode,
        local_only=local_only,
        investigation=investigation,
    )
    effective_core_mode = (
        "azure_sql"
        if not local_only and core_mode == "sqlite_shadow"
        else core_mode
    )
    persistence_patch = (
        patch.object(run_recon, "persist_sqlserver_run", side_effect=harness.persist)
        if effective_core_mode in {"sqlserver", "azure_sql"}
        else patch.object(run_recon, "persist_shadow_run", side_effect=harness.persist)
    )
    with (
        patch.object(run_recon, "capability_manifest", return_value=_all_capabilities()),
        patch.object(
            run_recon,
            "_verify_download_receipt",
            return_value={
                "source_item_id": "source-item",
                "space": "upload",
                "index_sha256": sha256_file(source_index),
                "_verified_index_path": str(source_index.resolve()),
            },
        ),
        patch.object(run_recon, "_run_command", side_effect=harness.command),
        persistence_patch,
        patch.dict(
            os.environ,
            {
                "NEXON_RECON_CORE_MODE": effective_core_mode,
                "NEXON_RECON_CORE_DSN": str(root / "core.db"),
            },
            clear=False,
        ),
    ):
        result = run_recon.run(args)
    return result, Path(result["run_root"])


def _resume_args(
    run_root: Path,
    *,
    investigation: Path | None = None,
    publication_receipt: Path | None = None,
    local_only: bool = False,
) -> Namespace:
    return _args(
        resume_root=run_root,
        investigation=investigation,
        publication_receipt=publication_receipt,
        local_only=local_only,
    )


def _resume_run(args: Namespace) -> dict:
    with patch.object(run_recon, "_verify_published_item"):
        return run_recon.resume_run(args)


def _publication_receipt(run_root: Path) -> dict:
    publication_set = read_json(run_root / "manifest" / "publication_set.json")
    run_prefix = (
        f"/recon-result-space/AAPT/{run_root.parent.parent.name}/"
        f"{run_root.parent.name}/{run_root.name}/"
    )
    uploaded = [
        {
            "local_path": item["local_path"],
            "sha256": item["sha256"],
            "item_id": f"item-{index}",
            "sharepoint_url": (
                "https://nexonap.sharepoint.com/sites/"
                "NexonReconciliationAutomation/Shared%20Documents"
                f"{run_prefix}{item['relative_path']}"
            ),
        }
        for index, item in enumerate(publication_set["artifacts"], start=1)
    ]
    run_manifest = read_json(run_root / "manifest" / "run_manifest.json")
    source_name = Path(run_manifest["source_file"]).name
    return {
        "run_id": run_root.name,
        "uploaded_artifacts": uploaded,
        "source_move_receipt": {
            "status": "moved",
            "item_id": "source-item",
            "sha256": run_manifest["source_checksum_sha256"],
            "sharepoint_url": (
                "https://nexonap.sharepoint.com/sites/"
                "NexonReconciliationAutomation/Shared%20Documents"
                f"{run_prefix}source/{source_name}"
            ),
        },
    }


class RunReconRemainingBranchTests(unittest.TestCase):
    def test_logical_run_path_contract(self) -> None:
        run_root = (
            Path("C:/sandbox/results")
            / "AAPT"
            / "2026"
            / "07"
            / "AAPT_20260723_201500_A1B2C"
        )
        self.assertEqual(
            "/recon-result-space/AAPT/2026/07/AAPT_20260723_201500_A1B2C",
            run_recon._logical_run_path("AAPT", run_root),
        )

    def test_command_and_provider_api_provenance_boundaries(self) -> None:
        success = subprocess.CompletedProcess(["python", "tool.py"], 0, "ok", "")
        failure = subprocess.CompletedProcess(["python", "tool.py"], 2, "", "bad input")
        with patch.object(run_recon.subprocess, "run", return_value=success):
            run_recon._run_command(["python", "tool.py"])
        with patch.object(run_recon.subprocess, "run", return_value=failure):
            with self.assertRaisesRegex(RuntimeError, r"tool_failed\[tool.py\]: bad input"):
                run_recon._run_command(["python", "tool.py"])

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invoice.pdf"
            source.write_bytes(b"invoice")
            manifest_path = root / "provider-api.json"
            valid_manifest = {
                "provider": "Equinix",
                "account_id": "ACC-1",
                "document_id": "DOC-1",
                "invoice_id": "INV-1",
                "output_file": str(source.resolve()),
                "output_sha256": sha256_file(source),
            }
            write_json(manifest_path, valid_manifest)

            def args(**updates: object) -> Namespace:
                values = {
                    "provider_api_manifest": manifest_path,
                    "provider": "Equinix",
                    "provider_api_account_id": "ACC-1",
                    "provider_api_document_id": "DOC-1",
                    "source_file": source,
                    "source_identity": "INV-1",
                }
                values.update(updates)
                return Namespace(**values)

            run_recon._validate_provider_api_provenance(args())
            cases = [
                ({"provider_api_manifest": None}, "manifest is required"),
                ({"provider": "AAPT"}, "manifest provider"),
                ({"provider_api_account_id": ""}, "provider account"),
                ({"provider_api_document_id": "OTHER"}, "provider document"),
                ({"source_file": None}, "output file"),
            ]
            for updates, message in cases:
                with self.subTest(message=message), self.assertRaisesRegex(RuntimeError, message):
                    run_recon._validate_provider_api_provenance(args(**updates))

            bad_manifest = dict(valid_manifest)
            bad_manifest["output_sha256"] = "0" * 64
            write_json(manifest_path, bad_manifest)
            with self.assertRaisesRegex(RuntimeError, "source checksum"):
                run_recon._validate_provider_api_provenance(args())
            write_json(manifest_path, valid_manifest)
            with self.assertRaisesRegex(RuntimeError, "source-identity"):
                run_recon._validate_provider_api_provenance(
                    args(source_identity="OTHER")
                )

    def test_investigation_contract_rejects_incomplete_or_unsafe_updates(self) -> None:
        matched = {
            "line_id": "matched",
            "ReconMatchStatus": "Matched",
        }
        unresolved = {
            "line_id": "unresolved",
            "ReconMatchStatus": "Not Matched",
        }
        valid_update = {
            "line_id": "unresolved",
            "agent_match_status": "no_match",
            "agent_evidence_summary": "No supported candidate.",
        }
        result = run_recon._apply_investigation(
            [matched, unresolved],
            {"run_id": "run-1", "rows": [valid_update]},
            expected_run_id="run-1",
        )
        self.assertEqual("Matched", result[0]["ReconMatchStatus"])
        self.assertTrue(result[1]["agent_review_required"])

        cases = [
            (
                {"run_id": "wrong", "rows": [valid_update]},
                "run_id does not match",
            ),
            (
                {"run_id": "run-1", "rows": {}},
                "expected rows list",
            ),
            (
                {"run_id": "run-1", "rows": [{"line_id": ""}]},
                "missing or duplicated",
            ),
            (
                {
                    "run_id": "run-1",
                    "rows": [valid_update, dict(valid_update)],
                },
                "missing or duplicated",
            ),
            (
                {"run_id": "run-1", "rows": []},
                "cover exactly",
            ),
            (
                {
                    "run_id": "run-1",
                    "rows": [{**valid_update, "provider": "Optus"}],
                },
                "disallowed fields",
            ),
            (
                {
                    "run_id": "run-1",
                    "rows": [{**valid_update, "agent_evidence_summary": " "}],
                },
                "evidence summary",
            ),
            (
                {
                    "run_id": "run-1",
                    "rows": [{**valid_update, "agent_match_status": "unsupported"}],
                },
                "unsupported agent_match_status",
            ),
        ]
        for investigation, message in cases:
            with self.subTest(message=message), self.assertRaisesRegex(RuntimeError, message):
                run_recon._apply_investigation(
                    [unresolved],
                    investigation,
                    expected_run_id="run-1",
                )

    def test_run_input_guards_provider_api_and_capability_fail_closed(self) -> None:
        parser_args = _args(run_mode="parser_validation")
        parser_args.copy = False
        with self.assertRaisesRegex(RuntimeError, "parser_validation requires --copy"):
            run_recon.run(parser_args)

        parser_api_args = _args(run_mode="parser_validation")
        parser_api_args.intake_mode = "provider_api"
        with self.assertRaisesRegex(
            RuntimeError, "parser_validation supports manual_upload"
        ):
            run_recon.run(parser_api_args)

        recon_args = _args(run_mode="reconciliation")
        recon_args.copy = True
        with self.assertRaisesRegex(RuntimeError, "reconciliation must move"):
            run_recon.run(recon_args)

        unavailable_api_args = _args(run_mode="reconciliation")
        unavailable_api_args.intake_mode = "provider_api"
        with self.assertRaisesRegex(RuntimeError, "provider_api_not_available"):
            run_recon.run(unavailable_api_args)

        available_api_args = _args(run_mode="reconciliation")
        available_api_args.intake_mode = "provider_api"
        enabled_api_config = {
            "features": {
                "db_update_enabled": False,
                "provider_api_enabled": True,
            },
            "billing": {"audit_required": True},
            "provider_api_adapters": {"aapt": True},
        }
        with (
            patch.object(run_recon, "load_config", return_value=enabled_api_config),
            patch.object(run_recon, "capability_manifest", return_value=_all_capabilities()),
            patch.object(run_recon, "_validate_provider_api_provenance") as provenance,
            self.assertRaisesRegex(RuntimeError, "run_input_missing"),
        ):
            run_recon.run(available_api_args)
        provenance.assert_called_once_with(available_api_args)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invoice.csv"
            source.write_text("invoice", encoding="utf-8")
            args = _args(
                source=source,
                result_root=root / "results",
                run_mode="parser_validation",
                local_only=False,
            )
            args.source_download_receipt = None
            with (
                patch.object(
                    run_recon,
                    "capability_manifest",
                    return_value=_all_capabilities(),
                ),
                self.assertRaisesRegex(
                    RuntimeError, "source_download_receipt_required"
                ),
            ):
                run_recon.run(args)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invoice.pdf"
            source.write_bytes(b"invoice")
            args = _args(
                source=source,
                result_root=root / "results",
                run_mode="parser_validation",
            )
            config = {
                "features": {
                    "db_update_enabled": False,
                },
                "billing": {"audit_required": True},
            }
            with (
                patch.object(run_recon, "load_config", return_value=config),
                patch.object(
                    run_recon,
                    "capability_manifest",
                    return_value={"capabilities": {"provider_parsing": False}},
                ),
                self.assertRaisesRegex(RuntimeError, "missing capabilities"),
            ):
                run_recon.run(args)

        missing_args = _args()
        with (
            patch.object(run_recon, "capability_manifest", return_value=_all_capabilities()),
            self.assertRaisesRegex(RuntimeError, "run_input_missing"),
        ):
            run_recon.run(missing_args)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invoice.csv"
            source.write_text("invoice", encoding="utf-8")
            args = _args(
                source=source,
                result_root=root / "results",
                run_mode="parser_validation",
                local_only=False,
            )
            args.sharepoint_binding = None
            with (
                patch.object(
                    run_recon,
                    "capability_manifest",
                    return_value=_all_capabilities(),
                ),
                self.assertRaisesRegex(RuntimeError, "sharepoint_binding_required"),
            ):
                run_recon.run(args)

    def test_ambiguous_staged_source_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invoice.csv"
            source.write_text("invoice", encoding="utf-8")
            result_root = root / "results"
            (result_root / "AAPT").mkdir(parents=True)
            args = _args(
                source=source,
                result_root=result_root,
                run_mode="parser_validation",
            )
            real_create_run = run_recon.create_run

            def create_ambiguous_run(**kwargs: object) -> Path:
                run_root = real_create_run(**kwargs)
                (run_root / "source" / "second.csv").write_text(
                    "second",
                    encoding="utf-8",
                )
                return run_root

            with (
                patch.object(run_recon, "capability_manifest", return_value=_all_capabilities()),
                patch.object(
                    run_recon,
                    "create_run",
                    side_effect=create_ambiguous_run,
                ),
                self.assertRaisesRegex(RuntimeError, "source_ambiguous"),
            ):
                run_recon.run(args)

    def test_archive_persistence_and_inline_investigation_lifecycle_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            result, _run_root = _new_run(
                Path(tmp),
                run_mode="parser_validation",
                source_suffix=".zip",
            )
            self.assertEqual("passed", result["validation"])

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invoice.zip"
            source.write_bytes(b"not-a-real-zip")
            result_root = root / "results"
            (result_root / "AAPT").mkdir(parents=True)
            args = _args(
                source=source,
                result_root=result_root,
                run_mode="parser_validation",
            )
            with (
                patch.object(run_recon, "capability_manifest", return_value=_all_capabilities()),
                patch.object(
                    run_recon,
                    "extract_zip",
                    return_value={
                        "archive": str(source),
                        "members": [],
                        "blocked": [{"reason": "unsafe"}],
                    },
                ),
                self.assertRaisesRegex(RuntimeError, "unsafe_archive"),
            ):
                run_recon.run(args)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invoice.csv"
            source.write_text("invoice", encoding="utf-8")
            result_root = root / "results"
            (result_root / "AAPT").mkdir(parents=True)
            args = _args(
                source=source,
                result_root=result_root,
                run_mode="reconciliation",
            )
            args.billing_sql_file = None
            harness = RuntimeHarness(False)
            with (
                patch.object(run_recon, "capability_manifest", return_value=_all_capabilities()),
                patch.object(run_recon, "_run_command", side_effect=harness.command),
                self.assertRaisesRegex(RuntimeError, "reconciliation_input_missing"),
            ):
                run_recon.run(args)

        for core_mode in ("sqlserver", "azure_sql"):
            with self.subTest(core_mode=core_mode), tempfile.TemporaryDirectory() as tmp:
                result, _run_root = _new_run(
                    Path(tmp),
                    run_mode="reconciliation",
                    matched=True,
                    core_mode=core_mode,
                )
                self.assertEqual("passed", result["validation"])

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invoice.csv"
            source.write_text("invoice", encoding="utf-8")
            billing_sql = root / "billing.sql"
            billing_sql.write_text("SELECT 1", encoding="utf-8")
            result_root = root / "results"
            (result_root / "AAPT").mkdir(parents=True)
            harness = RuntimeHarness(True)
            args = _args(
                source=source,
                result_root=result_root,
                run_mode="reconciliation",
            )
            with (
                patch.object(run_recon, "capability_manifest", return_value=_all_capabilities()),
                patch.object(run_recon, "_run_command", side_effect=harness.command),
                patch.object(run_recon, "persist_shadow_run", side_effect=harness.persist),
                patch.dict(
                    os.environ,
                    {
                        "NEXON_RECON_CORE_MODE": "sqlite_shadow",
                        "NEXON_RECON_CORE_DSN": "dsn",
                    },
                    clear=False,
                ),
                self.assertRaisesRegex(RuntimeError, "sqlite_shadow_not_allowed"),
            ):
                args.local_only = False
                args.sharepoint_binding = _sharepoint_binding(root)
                args.source_download_receipt = _download_receipt(root)
                with patch.object(
                    run_recon,
                    "_verify_download_receipt",
                    return_value={"source_item_id": "source-item"},
                ):
                    run_recon.run(args)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invoice.csv"
            source.write_text("invoice", encoding="utf-8")
            billing_sql = root / "billing.sql"
            billing_sql.write_text("SELECT 1", encoding="utf-8")
            result_root = root / "results"
            (result_root / "AAPT").mkdir(parents=True)
            harness = RuntimeHarness(True)
            args = _args(
                source=source,
                result_root=result_root,
                run_mode="reconciliation",
            )
            with (
                patch.object(run_recon, "capability_manifest", return_value=_all_capabilities()),
                patch.object(run_recon, "_run_command", side_effect=harness.command),
                patch.dict(
                    os.environ,
                    {
                        "NEXON_RECON_CORE_MODE": "unsupported",
                        "NEXON_RECON_CORE_DSN": "dsn",
                    },
                    clear=False,
                ),
                self.assertRaisesRegex(RuntimeError, "core_persistence_not_available"),
            ):
                run_recon.run(args)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            investigation = root / "investigation.json"
            write_json(investigation, {})
            harness = RuntimeHarness(False)
            source = root / "invoice.csv"
            source.write_text("invoice", encoding="utf-8")
            (root / "billing.sql").write_text("SELECT 1", encoding="utf-8")
            result_root = root / "results"
            (result_root / "AAPT").mkdir(parents=True)
            args = _args(
                source=source,
                result_root=result_root,
                run_mode="reconciliation",
                investigation=investigation,
            )
            with (
                patch.object(run_recon, "capability_manifest", return_value=_all_capabilities()),
                patch.object(run_recon, "_run_command", side_effect=harness.command),
                patch.object(run_recon, "persist_shadow_run", side_effect=harness.persist),
                patch.object(
                    run_recon,
                    "_apply_investigation",
                    side_effect=lambda rows, *_args, **_kwargs: [
                        {
                            **row,
                            "agent_match_status": "no_match",
                            "agent_evidence_summary": "No supported candidate.",
                            "agent_review_required": True,
                        }
                        for row in rows
                    ],
                ),
                patch.dict(
                    os.environ,
                    {
                        "NEXON_RECON_CORE_MODE": "sqlite_shadow",
                        "NEXON_RECON_CORE_DSN": str(root / "core.db"),
                    },
                    clear=False,
                ),
            ):
                result = run_recon.run(args)
            self.assertEqual("passed", result["validation"])
            self.assertTrue(
                (Path(result["run_root"]) / "refined-recon-report" / "refined-reconciliation.xlsx").is_file()
            )

    def test_resume_guards_and_nonlocal_investigation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            result, run_root = _new_run(
                root,
                run_mode="reconciliation",
                matched=False,
                local_only=False,
            )
            self.assertEqual("awaiting_exception_investigation", result["status"])
            with self.assertRaisesRegex(RuntimeError, "--investigation is required"):
                _resume_run(_resume_args(run_root))

            investigation = root / "investigation.json"
            write_json(
                investigation,
                {
                    "run_id": run_root.name,
                    "rows": [
                        {
                            "line_id": "line-1",
                            "agent_match_status": "no_match",
                            "agent_match_rule": "no_supported_candidate",
                            "agent_evidence_summary": "No supported candidate.",
                        }
                    ],
                },
            )
            resumed = _resume_run(
                _resume_args(
                    run_root,
                    investigation=investigation,
                    local_only=False,
                )
            )
            self.assertEqual("awaiting_publication", resumed["status"])

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            result, run_root = _new_run(
                root,
                run_mode="parser_validation",
            )
            state_path = run_root / "manifest" / "run_state.json"
            state = read_json(state_path)
            state["run_status"] = "running"
            write_json(state_path, state)
            with self.assertRaisesRegex(RuntimeError, "not awaiting investigation or publication"):
                _resume_run(_resume_args(run_root))

            state["run_status"] = "completed"
            write_json(state_path, state)
            with self.assertRaisesRegex(RuntimeError, "must be in running state"):
                _resume_run(_resume_args(run_root))

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            result, run_root = _new_run(
                root,
                run_mode="reconciliation",
                matched=False,
                local_only=True,
            )
            self.assertEqual("awaiting_exception_investigation", result["status"])
            investigation = root / "local-investigation.json"
            write_json(
                investigation,
                {
                    "run_id": run_root.name,
                    "rows": [
                        {
                            "line_id": "line-1",
                            "agent_match_status": "no_match",
                            "agent_evidence_summary": "No supported candidate.",
                        }
                    ],
                },
            )
            resumed = run_recon.run(
                _resume_args(
                    run_root,
                    investigation=investigation,
                    local_only=True,
                )
            )
            self.assertEqual("passed", resumed["validation"])

    def test_publication_receipt_rejection_matrix_and_provider_api_success(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            result, run_root = _new_run(
                root,
                run_mode="reconciliation",
                matched=True,
                local_only=False,
            )
            self.assertEqual("awaiting_publication", result["status"])
            receipt_path = root / "publication-receipt.json"
            valid = _publication_receipt(run_root)

            with self.assertRaisesRegex(RuntimeError, "--publication-receipt is required"):
                _resume_run(_resume_args(run_root))

            cases = [
                (
                    lambda payload: payload.update({"run_id": "wrong"}),
                    "receipt run_id",
                ),
                (
                    lambda payload: payload.update({"uploaded_artifacts": []}),
                    "requires uploaded_artifacts",
                ),
                (
                    lambda payload: payload["uploaded_artifacts"][0].update(
                        {"sharepoint_url": "http://example.invalid/wrong"}
                    ),
                    "artifact URL/item ID",
                ),
                (
                    lambda payload: payload["uploaded_artifacts"][0].update(
                        {"sha256": "0" * 64}
                    ),
                    "paths/checksums",
                ),
                (
                    lambda payload: payload["uploaded_artifacts"].append(
                        copy.deepcopy(payload["uploaded_artifacts"][0])
                    ),
                    "paths/checksums",
                ),
                (
                    lambda payload: payload.update({"source_move_receipt": {}}),
                    "source move receipt",
                ),
                (
                    lambda payload: payload["source_move_receipt"].update(
                        {"item_id": "different-source-item"}
                    ),
                    "source move receipt",
                ),
            ]
            for mutation, message in cases:
                payload = copy.deepcopy(valid)
                mutation(payload)
                write_json(receipt_path, payload)
                with self.subTest(message=message), self.assertRaisesRegex(RuntimeError, message):
                    _resume_run(
                        _resume_args(run_root, publication_receipt=receipt_path)
                    )

            publication_set_path = run_root / "manifest" / "publication_set.json"
            original_set = read_json(publication_set_path)
            for mutation, message in (
                (
                    lambda payload: payload.update({"run_id": "wrong"}),
                    "frozen publication set has wrong",
                ),
                (
                    lambda payload: payload.update({"artifacts": []}),
                    "frozen publication set is empty",
                ),
            ):
                payload = copy.deepcopy(original_set)
                mutation(payload)
                write_json(publication_set_path, payload)
                write_json(receipt_path, valid)
                with self.subTest(message=message), self.assertRaisesRegex(RuntimeError, message):
                    _resume_run(
                        _resume_args(run_root, publication_receipt=receipt_path)
                    )
            write_json(publication_set_path, original_set)

            artifact_path = Path(original_set["artifacts"][0]["local_path"])
            original_artifact = artifact_path.read_bytes()
            try:
                artifact_path.write_bytes(original_artifact + b"tampered")
                write_json(receipt_path, valid)
                with self.assertRaisesRegex(RuntimeError, "frozen artifact changed"):
                    _resume_run(
                        _resume_args(run_root, publication_receipt=receipt_path)
                    )
            finally:
                artifact_path.write_bytes(original_artifact)

            run_manifest_path = run_root / "manifest" / "run_manifest.json"
            run_manifest = read_json(run_manifest_path)
            run_manifest["intake_mode"] = "provider_api"
            write_json(run_manifest_path, run_manifest)
            provider_api_receipt = copy.deepcopy(valid)
            provider_api_receipt.pop("source_move_receipt")
            new_manifest_hash = sha256_file(run_manifest_path)
            for item in original_set["artifacts"]:
                if Path(item["local_path"]) == run_manifest_path.resolve():
                    item["sha256"] = new_manifest_hash
            for item in provider_api_receipt["uploaded_artifacts"]:
                if Path(item["local_path"]) == run_manifest_path.resolve():
                    item["sha256"] = new_manifest_hash
            write_json(publication_set_path, original_set)
            write_json(receipt_path, provider_api_receipt)
            resumed = _resume_run(
                _resume_args(run_root, publication_receipt=receipt_path)
            )
            self.assertEqual("passed", resumed["validation"])

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            result, run_root = _new_run(
                root,
                run_mode="reconciliation",
                matched=True,
                local_only=False,
            )
            self.assertEqual("awaiting_publication", result["status"])
            receipt_path = root / "manual-publication-receipt.json"
            write_json(receipt_path, _publication_receipt(run_root))
            resumed = _resume_run(
                _resume_args(run_root, publication_receipt=receipt_path)
            )
            self.assertEqual("passed", resumed["validation"])

    def test_cli_output_and_console_paths(self) -> None:
        result = {"validation": "passed"}
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "result.json"
            with (
                patch.object(sys, "argv", ["run_recon.py", "--output", str(output)]),
                patch.object(run_recon, "run", return_value=result),
            ):
                self.assertEqual(0, run_recon.main())
            self.assertEqual(result, read_json(output))

            with (
                patch.object(sys, "argv", ["run_recon.py"]),
                patch.object(run_recon, "run", return_value=result),
            ):
                self.assertEqual(0, run_recon.main())


class ValidateRunRemainingBranchTests(unittest.TestCase):
    def _valid_parser_run(self, root: Path) -> Path:
        _result, run_root = _new_run(root, run_mode="parser_validation")
        return run_root

    def _valid_reconciliation_run(self, root: Path) -> Path:
        _result, run_root = _new_run(
            root,
            run_mode="reconciliation",
            matched=True,
        )
        persisted = read_json(
            run_root / "normalized" / "persisted_match_results.json"
        ).get("rows", [])
        run_recon.write_reports(
            raw_rows=persisted,
            refined_input_rows=persisted,
            raw_output=run_root / "raw-recon-report" / "raw-reconciliation.xlsx",
            refined_output=(
                run_root
                / "refined-recon-report"
                / "refined-reconciliation.xlsx"
            ),
            manifest=run_root / "manifest" / "report_manifest.json",
            config=validate_run.load_config(CONFIG),
            run_path=run_recon._logical_run_path("AAPT", run_root),
            period="2026-06",
        )
        return run_root

    def _mutate_json_failure(
        self,
        run_root: Path,
        relative_path: str,
        mutation,
        message: str,
        *,
        run_mode: str = "parser_validation",
    ) -> None:
        path = run_root / relative_path
        original = path.read_bytes()
        try:
            payload = read_json(path)
            mutation(payload)
            write_json(path, payload)
            with self.assertRaisesRegex(RuntimeError, message):
                validate_run.validate_run(
                    run_root,
                    validate_run.load_config(CONFIG),
                    run_mode,
                )
        finally:
            path.write_bytes(original)

    def test_workbook_loader_rows_metadata_and_path_shape_errors(self) -> None:
        real_import = builtins.__import__

        def fail_openpyxl(name: str, *args: object, **kwargs: object):
            if name == "openpyxl":
                raise ImportError("missing")
            return real_import(name, *args, **kwargs)

        with patch.object(builtins, "__import__", side_effect=fail_openpyxl):
            with self.assertRaisesRegex(RuntimeError, "openpyxl is required"):
                validate_run._load_workbook(Path("missing.xlsx"))

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            missing_result = root / "missing-result.xlsx"
            workbook = Workbook()
            workbook.active.title = "Other"
            workbook.save(missing_result)
            with self.assertRaisesRegex(RuntimeError, "missing Result sheet"):
                validate_run.workbook_rows(missing_result)

            empty_result = root / "empty-result.xlsx"
            workbook = Workbook()
            workbook.active.title = "Result"
            workbook.save(empty_result)
            with self.assertRaisesRegex(RuntimeError, "Result sheet is empty"):
                validate_run.workbook_rows(empty_result)

            run_root = (
                root
                / "AAPT"
                / "2026"
                / "07"
                / "AAPT_20260723_201500_A1B2C"
            )
            metadata = root / "metadata.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Do not change"
            sheet.append(
                ["RunPath", run_recon._logical_run_path("AAPT", run_root)]
            )
            sheet.append(["ReconciliationPeriod", "2026-05"])
            workbook.save(metadata)
            validate_run.assert_workbook_metadata(
                metadata,
                run_root,
                "2026-05",
                provider="AAPT",
            )
            with self.assertRaisesRegex(RuntimeError, "RunPath metadata"):
                validate_run.assert_workbook_metadata(
                    metadata,
                    run_root.parent / "AAPT_20260723_201501_A1B2D",
                    "2026-05",
                    provider="AAPT",
                )
            with self.assertRaisesRegex(RuntimeError, "reconciliation period"):
                validate_run.assert_workbook_metadata(
                    metadata,
                    run_root,
                    "2026-06",
                    provider="AAPT",
                )
            workbook = validate_run._load_workbook(metadata)
            workbook["Do not change"]["B1"] = str(run_root)
            workbook.save(metadata)
            validate_run.assert_workbook_metadata(metadata, run_root, "2026-05")

            bad_run_root = root / "Optus" / "2026" / "07" / "AAPT_20260709_153012_A1B2C"
            with self.assertRaisesRegex(RuntimeError, "Run path must end"):
                validate_run.assert_result_path_shape(
                    bad_run_root,
                    "AAPT_20260709_153012_A1B2C",
                )

    def test_secret_scanning_and_evidence_summary_boundaries(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            validate_run.assert_no_secret_markers(root)

            logs = root / "logs"
            logs.mkdir()
            (logs / "directory").mkdir()
            validate_run.assert_no_secret_markers(root)

            bad_workbook = logs / "bad.xlsx"
            bad_workbook.write_bytes(b"not a workbook")
            with self.assertRaisesRegex(RuntimeError, "Unable to scan output"):
                validate_run.assert_no_secret_markers(root)
            bad_workbook.unlink()

            secret = logs / "secret.txt"
            secret.write_text("client_secret=do-not-log", encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "Secret-like marker"):
                validate_run.assert_no_secret_markers(root)

        with self.assertRaisesRegex(RuntimeError, "evidence_summary is required"):
            validate_run.assert_evidence_summary(
                {"agent_match_status": "no_match", "agent_evidence_summary": ""},
                auto_matched_mode="blank",
                max_chars=240,
            )
        with self.assertRaisesRegex(RuntimeError, "must be a single line"):
            validate_run.assert_evidence_summary(
                {
                    "agent_match_status": "no_match",
                    "agent_evidence_summary": "line one\nline two",
                },
                auto_matched_mode="blank",
                max_chars=240,
            )

    def test_parser_and_audit_contract_rejection_matrix(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            run_root = self._valid_parser_run(Path(tmp))
            warnings_path = run_root / "logs" / "parser_warnings.json"
            original_warnings = warnings_path.read_bytes()
            try:
                warnings_path.unlink()
                with self.assertRaisesRegex(RuntimeError, "Missing parser artifact"):
                    validate_run.assert_parser_contract(run_root, run_root.name, "AAPT")
            finally:
                warnings_path.write_bytes(original_warnings)

            cases = [
                (
                    "logs/parser_warnings.json",
                    lambda _payload: {"not": "a list"},
                    "warnings artifact must be a list",
                ),
                (
                    "logs/parser_warnings.json",
                    lambda _payload: [{"severity": "error"}],
                    "Parser error warnings",
                ),
                (
                    "manifest/parser_manifest.json",
                    lambda payload: {**payload, "provider": "Optus"},
                    "manifest identity",
                ),
                (
                    "normalized/provider_lines.json",
                    lambda payload: {**payload, "lines": {}},
                    "must contain invoice_headers and lines",
                ),
                (
                    "manifest/parser_manifest.json",
                    lambda payload: {**payload, "accounting_complete": False},
                    "complete member/row accounting",
                ),
                (
                    "manifest/parser_manifest.json",
                    lambda payload: {**payload, "source_rows": 2},
                    "source row accounting",
                ),
                (
                    "manifest/parser_manifest.json",
                    lambda payload: {**payload, "parsed_rows": 2},
                    "parsed_rows",
                ),
                (
                    "normalized/provider_lines.json",
                    lambda payload: {
                        **payload,
                        "lines": [{**payload["lines"][0], "line_id": ""}],
                    },
                    "requires line_id",
                ),
                (
                    "normalized/provider_lines.json",
                    lambda payload: {
                        **payload,
                        "lines": [{**payload["lines"][0], "provider": "Optus"}],
                    },
                    "not bound to the current run/provider",
                ),
            ]
            for relative_path, transform, message in cases:
                path = run_root / relative_path
                original = path.read_bytes()
                try:
                    write_json(path, transform(read_json(path)))
                    with self.subTest(message=message), self.assertRaisesRegex(RuntimeError, message):
                        validate_run.assert_parser_contract(
                            run_root,
                            run_root.name,
                            "AAPT",
                        )
                finally:
                    path.write_bytes(original)

            with self.assertRaisesRegex(RuntimeError, "Missing parser warnings artifact"):
                warnings_path.unlink()
                validate_run.assert_parser_warnings_resolved(run_root)
            warnings_path.write_bytes(original_warnings)
            write_json(warnings_path, {"bad": True})
            with self.assertRaisesRegex(RuntimeError, "must be a list"):
                validate_run.assert_parser_warnings_resolved(run_root)
            write_json(warnings_path, [{"severity": "error"}])
            with self.assertRaisesRegex(RuntimeError, "Parser error warnings"):
                validate_run.assert_parser_warnings_resolved(run_root)
            warnings_path.write_bytes(original_warnings)
            validate_run.assert_parser_warnings_resolved(run_root)

            audit_path = run_root / "manifest" / "audit_manifest.json"
            state_path = run_root / "manifest" / "run_state.json"
            original_audit = audit_path.read_bytes()
            original_state = state_path.read_bytes()
            try:
                audit_path.unlink()
                with self.assertRaisesRegex(RuntimeError, "audit manifest is missing"):
                    validate_run.assert_audit(run_root, run_root.name)
                audit_path.write_bytes(original_audit)
                state_path.unlink()
                with self.assertRaisesRegex(RuntimeError, "run state is missing"):
                    validate_run.assert_audit(run_root, run_root.name)
                state_path.write_bytes(original_state)

                audit = read_json(audit_path)
                audit["run_id"] = "wrong"
                write_json(audit_path, audit)
                with self.assertRaisesRegex(RuntimeError, "run identity"):
                    validate_run.assert_audit(run_root, run_root.name)
                audit_path.write_bytes(original_audit)

                audit = read_json(audit_path)
                audit["accepted_resolution_update_attempted"] = True
                write_json(audit_path, audit)
                with self.assertRaisesRegex(RuntimeError, "update was attempted"):
                    validate_run.assert_audit(run_root, run_root.name)
                audit_path.write_bytes(original_audit)

                state = read_json(state_path)
                state["stages"] = []
                write_json(state_path, state)
                with self.assertRaisesRegex(RuntimeError, "no stage ledger"):
                    validate_run.assert_audit(run_root, run_root.name)
            finally:
                audit_path.write_bytes(original_audit)
                state_path.write_bytes(original_state)

    def test_reconciliation_runtime_and_workbook_rejection_matrix(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            run_root = self._valid_reconciliation_run(Path(tmp))
            config = validate_run.load_config(CONFIG)
            state = read_json(run_root / "manifest" / "run_state.json")

            candidates_path = run_root / "evidence" / "billing_candidates.json"
            original_candidates = candidates_path.read_bytes()
            try:
                candidates_path.unlink()
                with self.assertRaisesRegex(RuntimeError, "runtime artifacts are missing"):
                    validate_run.assert_reconciliation_runtime(run_root, state, 1)
            finally:
                candidates_path.write_bytes(original_candidates)

            mutations = [
                (
                    "manifest/persistence_manifest.json",
                    lambda payload: payload.update({"provider": "Optus"}),
                    "Persistence manifest identity",
                ),
                (
                    "manifest/persistence_manifest.json",
                    lambda payload: payload.update({"transaction": "rolled_back"}),
                    "committed transaction",
                ),
                (
                    "manifest/persistence_manifest.json",
                    lambda payload: payload.update({"supplier_line_count": 2}),
                    "supplier-line count",
                ),
                (
                    "manifest/persistence_manifest.json",
                    lambda payload: payload.update({"result_count": 2}),
                    "result count",
                ),
                (
                    "manifest/persistence_manifest.json",
                    lambda payload: payload.update({"result_count": 0}),
                    "fewer results",
                ),
                (
                    "logs/billing_query_log.json",
                    lambda payload: payload.clear(),
                    "at least one audited query chunk",
                ),
                (
                    "evidence/billing_candidates.json",
                    lambda payload: payload.update({"provider": "Optus"}),
                    "Billing candidate evidence identity",
                ),
                (
                    "manifest/audit_manifest.json",
                    lambda payload: payload.update({"query_logs": []}),
                    "query-log provenance",
                ),
                (
                    "manifest/audit_manifest.json",
                    lambda payload: payload["query_logs"][0].update({"sha256": "0" * 64}),
                    "query log hash",
                ),
                (
                    "normalized/match_results.json",
                    lambda payload: payload["rows"][0].update({"line_id": "other"}),
                    "line identities",
                ),
                (
                    "normalized/provider_lines.json",
                    lambda payload: payload["lines"][0].update({"line_id": "other"}),
                    "line identities",
                ),
                (
                    "manifest/run_state.json",
                    lambda payload: payload["stages"]["raw_workbook"].update(
                        {"status": "pending"}
                    ),
                    "Required reconciliation stages",
                ),
                (
                    "manifest/run_state.json",
                    lambda payload: payload["stages"]["notification"].update(
                        {"status": "running"}
                    ),
                    "stage is not terminal",
                ),
            ]
            for relative_path, mutation, message in mutations:
                self._mutate_json_failure(
                    run_root,
                    relative_path,
                    mutation,
                    message,
                    run_mode="reconciliation",
                )

            state_path = run_root / "manifest" / "run_state.json"
            original_state = state_path.read_bytes()
            try:
                state = read_json(state_path)
                state["stages"]["publication"]["status"] = "completed"
                write_json(state_path, state)
                with self.assertRaisesRegex(RuntimeError, "publication has no receipt"):
                    validate_run.assert_reconciliation_runtime(run_root, state, 1)
            finally:
                state_path.write_bytes(original_state)

            manifest_path = run_root / "manifest" / "report_manifest.json"
            raw_path = run_root / "raw-recon-report" / "raw-reconciliation.xlsx"
            refined_path = run_root / "refined-recon-report" / "refined-reconciliation.xlsx"
            original_manifest = manifest_path.read_bytes()
            original_raw = raw_path.read_bytes()
            original_refined = refined_path.read_bytes()
            try:
                manifest = read_json(manifest_path)
                manifest["raw_output"] = str(
                    run_root / "raw-recon-report" / "other.xlsx"
                )
                with self.assertRaisesRegex(RuntimeError, "raw path"):
                    validate_run.validate_workbooks(run_root, config, manifest)

                raw_path.unlink()
                with self.assertRaisesRegex(RuntimeError, "Raw workbook is missing"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                raw_path.write_bytes(original_raw)

                workbook = validate_run._load_workbook(raw_path)
                workbook["Result"].cell(1, 1).value = "WrongHeader"
                workbook.save(raw_path)
                with self.assertRaisesRegex(RuntimeError, "35-column contract"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                raw_path.write_bytes(original_raw)

                workbook = validate_run._load_workbook(raw_path)
                workbook["Do not change"]["B2"] = "2025-01"
                workbook.save(raw_path)
                with self.assertRaisesRegex(RuntimeError, "reconciliation period"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                raw_path.write_bytes(original_raw)

                workbook = validate_run._load_workbook(raw_path)
                workbook.create_sheet("Extra")
                workbook.save(raw_path)
                with self.assertRaisesRegex(RuntimeError, "sheet contract mismatch"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                raw_path.write_bytes(original_raw)

                workbook = validate_run._load_workbook(raw_path)
                workbook["Result"].cell(2, 1).value = "tampered"
                workbook.save(raw_path)
                with self.assertRaisesRegex(RuntimeError, "persisted reconciliation results"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                raw_path.write_bytes(original_raw)

                refined_path.unlink()
                with self.assertRaisesRegex(RuntimeError, "Refined workbook is missing"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                refined_path.write_bytes(original_refined)

                manifest = read_json(manifest_path)
                manifest["refined_output"] = str(run_root / "refined-recon-report" / "other.xlsx")
                with self.assertRaisesRegex(RuntimeError, "refined path"):
                    validate_run.validate_workbooks(run_root, config, manifest)

                workbook = validate_run._load_workbook(refined_path)
                workbook["Result"].cell(1, 1).value = "WrongHeader"
                workbook.save(refined_path)
                with self.assertRaisesRegex(RuntimeError, "preserve raw fields"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                refined_path.write_bytes(original_refined)

                workbook = validate_run._load_workbook(refined_path)
                workbook.create_sheet("Extra")
                workbook.save(refined_path)
                with self.assertRaisesRegex(RuntimeError, "sheet contract mismatch"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                refined_path.write_bytes(original_refined)

                workbook = validate_run._load_workbook(refined_path)
                headers = [cell.value for cell in workbook["Result"][1]]
                agent_status_column = headers.index("agent_match_status") + 1
                workbook["Result"].cell(2, agent_status_column).value = "unsupported"
                workbook.save(refined_path)
                with self.assertRaisesRegex(RuntimeError, "Invalid agent_match_status"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                refined_path.write_bytes(original_refined)

                workbook = validate_run._load_workbook(refined_path)
                headers = [cell.value for cell in workbook["Result"][1]]
                human_status_column = headers.index("human_verified_status") + 1
                workbook["Result"].cell(2, human_status_column).value = "unsupported"
                workbook.save(refined_path)
                with self.assertRaisesRegex(RuntimeError, "Invalid human_verified_status"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                refined_path.write_bytes(original_refined)

                workbook = validate_run._load_workbook(refined_path)
                workbook["Result"].delete_rows(2)
                workbook.save(refined_path)
                with self.assertRaisesRegex(RuntimeError, "row counts differ"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                refined_path.write_bytes(original_refined)

                workbook = validate_run._load_workbook(refined_path)
                raw_column = 1
                workbook["Result"].cell(2, raw_column).value = "tampered"
                workbook.save(refined_path)
                with self.assertRaisesRegex(RuntimeError, "raw fields"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                refined_path.write_bytes(original_refined)

                workbook = validate_run._load_workbook(refined_path)
                headers = [
                    cell.value
                    for cell in workbook["Result"][1]
                ]
                invoice_column = headers.index("human_verified_invoice_number") + 1
                status_column = headers.index("human_verified_status") + 1
                workbook["Result"].cell(2, status_column).value = "verified"
                workbook["Result"].cell(2, invoice_column).value = ""
                workbook.save(refined_path)
                with self.assertRaisesRegex(RuntimeError, "human_verified_invoice_number"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                refined_path.write_bytes(original_refined)

                workbook = validate_run._load_workbook(refined_path)
                headers = [cell.value for cell in workbook["Result"][1]]
                summary_column = headers.index("agent_evidence_summary") + 1
                workbook["Result"].cell(2, summary_column).value = "x" * 241
                workbook.save(refined_path)
                with self.assertRaisesRegex(RuntimeError, "exceeds max_chars"):
                    validate_run.validate_workbooks(
                        run_root,
                        config,
                        read_json(manifest_path),
                    )
                refined_path.write_bytes(original_refined)

                manifest = read_json(manifest_path)
                manifest["row_count"] = 2
                with self.assertRaisesRegex(RuntimeError, "row_count"):
                    validate_run.validate_workbooks(run_root, config, manifest)
            finally:
                manifest_path.write_bytes(original_manifest)
                raw_path.write_bytes(original_raw)
                refined_path.write_bytes(original_refined)

    def test_defensive_policy_guards_final_count_and_validator_cli(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            run_root = self._valid_reconciliation_run(root)
            config = validate_run.load_config(CONFIG)
            report_manifest = read_json(run_root / "manifest" / "report_manifest.json")

            # The preceding exact-header check makes this guard unreachable with
            # today's fixed raw columns; simulate policy drift to verify it fails closed.
            with patch.object(
                validate_run,
                "workbook_rows",
                return_value=(
                    validate_run.RAW_WORKBOOK_COLUMNS[:-1] + ["agent_policy_probe"],
                    [],
                    ["Result", "Adjustment", "Do not change"],
                ),
            ), patch.object(
                validate_run,
                "RAW_WORKBOOK_COLUMNS",
                validate_run.RAW_WORKBOOK_COLUMNS[:-1] + ["agent_policy_probe"],
            ), patch.object(validate_run, "assert_workbook_metadata"):
                with self.assertRaisesRegex(RuntimeError, "Agent/human columns leaked"):
                    validate_run.validate_workbooks(run_root, config, report_manifest)

            # Approved and excluded columns are currently disjoint, so exercise the
            # defensive overlap guard by simulating a future policy-constant conflict.
            with patch.object(
                validate_run,
                "EXCLUDED_PHASE1_COLUMNS",
                ("agent_match_status",),
            ):
                with self.assertRaisesRegex(RuntimeError, "Excluded Phase 1 columns"):
                    validate_run.validate_workbooks(run_root, config, report_manifest)

            with (
                patch.object(validate_run, "validate_workbooks", return_value=0),
                self.assertRaisesRegex(RuntimeError, "Parsed and reported row counts differ"),
            ):
                validate_run.validate_run(run_root, config, "reconciliation")

            run_manifest_path = run_root / "manifest" / "run_manifest.json"
            original_run_manifest = run_manifest_path.read_bytes()
            try:
                run_manifest = read_json(run_manifest_path)
                run_manifest["run_id"] = "wrong"
                write_json(run_manifest_path, run_manifest)
                with self.assertRaisesRegex(RuntimeError, "run_id does not match"):
                    validate_run.validate_run(run_root, config, "reconciliation")
                run_manifest_path.write_bytes(original_run_manifest)

                run_manifest = read_json(run_manifest_path)
                run_manifest["db_update_enabled"] = True
                write_json(run_manifest_path, run_manifest)
                with self.assertRaisesRegex(RuntimeError, "update was not proven disabled"):
                    validate_run.validate_run(run_root, config, "reconciliation")
            finally:
                run_manifest_path.write_bytes(original_run_manifest)

            invalid_run_root = run_root.with_name("invalid-run-id")
            with self.assertRaisesRegex(RuntimeError, "Invalid run_id"):
                validate_run.validate_run(
                    invalid_run_root,
                    config,
                    "parser_validation",
                )

            missing_subdir = run_root / "extracted"
            missing_subdir.rmdir()
            try:
                with self.assertRaisesRegex(RuntimeError, "Missing run subdir"):
                    validate_run.validate_run(
                        run_root,
                        config,
                        "parser_validation",
                    )
            finally:
                missing_subdir.mkdir()

            run_manifest_path = run_root / "manifest" / "run_manifest.json"
            original_run_manifest = run_manifest_path.read_bytes()
            run_manifest_path.unlink()
            try:
                with self.assertRaisesRegex(RuntimeError, "Missing run manifest"):
                    validate_run.validate_run(
                        run_root,
                        config,
                        "parser_validation",
                    )
            finally:
                run_manifest_path.write_bytes(original_run_manifest)

            report_manifest_path = run_root / "manifest" / "report_manifest.json"
            original_report_manifest = report_manifest_path.read_bytes()
            report_manifest_path.unlink()
            try:
                with self.assertRaisesRegex(RuntimeError, "Missing report manifest"):
                    validate_run.validate_run(
                        run_root,
                        config,
                        "reconciliation",
                    )
            finally:
                report_manifest_path.write_bytes(original_report_manifest)

            output = root / "validation.json"
            with (
                patch.object(
                    sys,
                    "argv",
                    [
                        "validate_run.py",
                        "--config",
                        str(CONFIG),
                        "--run-root",
                        str(run_root),
                        "--run-mode",
                        "reconciliation",
                        "--output",
                        str(output),
                    ],
                ),
            ):
                self.assertEqual(0, validate_run.main())
            self.assertEqual("passed", read_json(output)["validation"])

            with (
                patch.object(
                    sys,
                    "argv",
                    [
                        "validate_run.py",
                        "--config",
                        str(CONFIG),
                        "--run-root",
                        str(run_root),
                        "--run-mode",
                        "reconciliation",
                    ],
                ),
            ):
                self.assertEqual(0, validate_run.main())


if __name__ == "__main__":
    unittest.main()

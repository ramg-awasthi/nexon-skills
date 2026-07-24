from __future__ import annotations

import builtins
import copy
import gc
import hashlib
import json
import os
import stat
import sys
import tempfile
import types
import unittest
from argparse import Namespace
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch
from urllib.error import HTTPError

from openpyxl import Workbook


SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS))

from recon_core import billing_query, common, core_persistence, intake_run  # noqa: E402
from recon_core import optional_db_update, provider_api_download, record_notification  # noqa: E402
from recon_core import run_state, safe_unpack, write_reports  # noqa: E402


VALID_SQL = (
    "SELECT Carrier_Name AS provider, AccountNumber AS provider_account, "
    "BillingDate AS transaction_date, ServiceNumber AS service_id "
    "FROM Finance.GenericNexonBilling"
)
RUN_ID = "AAPT_20260709_153012_A1B2C"


def call_main(module: object, argv: list[object]) -> int:
    old_argv = sys.argv[:]
    sys.argv = [getattr(module, "__name__", "tool")] + [str(value) for value in argv]
    try:
        return module.main()
    finally:
        sys.argv = old_argv


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def write_config(
    path: Path,
    *,
    candidate_limit: int = 20,
    db_update_enabled: bool = False,
) -> None:
    path.write_text(
        f"""
features:
  billing_query_enabled: true
  provider_api_enabled: true
  core_persistence_enabled: true
  db_update_enabled: {str(db_update_enabled).lower()}
provider_api_adapters:
  equinix: true
billing:
  mode: read_only_sql
  agent_sql_allowed: true
  audit_required: true
limits:
  investigation_rows_per_batch: 100
  billing_query_row_limit: 5000
  billing_query_timeout_seconds: 30
  investigation_candidates_per_row: {candidate_limit}
  investigation_query_rounds: 2
reports:
  evidence_summary:
    auto_matched: short
    max_chars: 160
db_update_policy:
  allow_deferred_without_invoice_number: true
""".lstrip(),
        encoding="utf-8",
    )


def normalized_payload() -> dict:
    return {
        "lines": [
            {
                "run_id": RUN_ID,
                "provider": "AAPT",
                "line_id": "line-1",
                "provider_account": "ACC-1",
                "service_id_raw": "SVC-1",
                "service_id_normalized": "SVC-1",
                "invoice_number": "INV-1",
                "billing_period_start": "2026-06-01",
                "billing_period_end": "2026-06-30",
            }
        ]
    }


def persistence_payloads() -> tuple[dict, dict, dict]:
    header = {
        "request_key": f"{RUN_ID}:invoice-1",
        "invoice_identity": "AAPT:INV-1",
        "invoice_number": "INV-1",
        "billing_period_start": "2026-06-01",
        "billing_period_end": "2026-06-30",
    }
    line = {
        "line_id": "line-1",
        "run_id": RUN_ID,
        "provider": "AAPT",
        "request_key": header["request_key"],
        "invoice_identity": header["invoice_identity"],
        "service_id_normalized": "SVC-1",
        "billing_period_start": "2026-06-01",
        "billing_period_end": "2026-06-30",
        "supplier_amount": "12.34",
    }
    candidate = {
        "candidate_id": "candidate-1",
        "invoice_number": "CUST-INV-1",
        "customer_name": "Customer",
        "account_number": "ACC-1",
        "service_id": "SVC-1",
        "billing_date": "2026-06-30",
        "amount_excl_gst": "12.34",
    }
    match = {
        "line_id": "line-1",
        "run_id": RUN_ID,
        "provider": "AAPT",
        "ReconMatchStatus": "Matched",
        "candidate_snapshot": candidate,
    }
    return (
        {"invoice_headers": [header], "lines": [line]},
        {
            "run_id": RUN_ID,
            "provider": "AAPT",
            "candidates_by_line": {"line-1": [candidate]},
        },
        {"rows": [match]},
    )


class AdditionalStrictBranchEdgeTests(unittest.TestCase):
    def test_billing_helpers_and_database_driver_edges(self) -> None:
        self.assertEqual(
            hashlib.sha256(b"{}\n").hexdigest(),
            billing_query._json_payload_hash({}),
        )
        self.assertEqual("AAPT", billing_query._line_params(normalized_payload()["lines"][0])["provider"])
        with self.assertRaisesRegex(ValueError, "empty billing chunk"):
            billing_query._chunk_params([])
        self.assertNotIn("empty", billing_query._param_hashes({"empty": "", "value": "x"}))
        with self.assertRaisesRegex(RuntimeError, "parameter_missing"):
            billing_query._qmark_sql("select :missing", {})
        query, values = billing_query._qmark_sql("select :a, :a, :b", {"a": 1, "b": 2})
        self.assertEqual("select ?, ?, ?", query)
        self.assertEqual([1, 1, 2], values)

        with patch.dict(os.environ, {"NEXON_RECON_BILLING_MODE": "postgres"}, clear=False):
            self.assertIn("jsonb_array_elements_text", billing_query._scoped_sql(VALID_SQL, "postgres"))
        with self.assertRaisesRegex(RuntimeError, "unapproved source columns"):
            billing_query._assert_read_only_sql(
                "SELECT provider, provider_account, transaction_date, wrong AS service_id "
                "FROM Finance.GenericNexonBilling"
            )
        with self.assertRaisesRegex(RuntimeError, "forbidden column identifiers"):
            billing_query._assert_read_only_sql(
                "SELECT provider, provider_account, transaction_date, service_id, secret_value "
                "FROM Finance.GenericNexonBilling"
            )
        with patch("sqlglot.parse_one", side_effect=ValueError("bad")):
            with self.assertRaisesRegex(RuntimeError, "could not be parsed"):
                billing_query._assert_read_only_sql(VALID_SQL)
        with patch("sqlglot.parse_one", return_value=types.SimpleNamespace(find_all=lambda *_: [], find=lambda *_: None)):
            with self.assertRaisesRegex(RuntimeError, "no SELECT projection"):
                billing_query._assert_read_only_sql(VALID_SQL)

        old_import = builtins.__import__

        def blocked_import(name: str, *args: object, **kwargs: object) -> object:
            if name in {"pyodbc", "psycopg", "sqlglot"}:
                raise ImportError(name)
            return old_import(name, *args, **kwargs)

        with patch.object(builtins, "__import__", side_effect=blocked_import):
            with self.assertRaisesRegex(RuntimeError, "sqlglot is required"):
                billing_query._assert_read_only_sql(VALID_SQL)
            with self.assertRaisesRegex(RuntimeError, "pyodbc is required"):
                billing_query._execute_sqlserver("dsn", "select 1", {}, timeout_seconds=1, row_limit=1)
            with self.assertRaisesRegex(RuntimeError, "psycopg is required"):
                billing_query._execute_postgres("dsn", "select 1", {}, timeout_seconds=1, row_limit=1)

        class FakeSqlServerCursor:
            def __init__(self, rows: list[tuple[str]]) -> None:
                self.rows = rows
                self.description = [("service_id",)]
                self.timeout = 0
                self.executed: list[tuple[object, ...]] = []

            def execute(self, *args: object) -> None:
                self.executed.append(args)

            def fetchmany(self, _count: int) -> list[tuple[str]]:
                return self.rows

        class FakeSqlServerConnection:
            def __init__(self, rows: list[tuple[str]]) -> None:
                self.cursor_value = FakeSqlServerCursor(rows)
                self.rolled_back = False
                self.closed = False

            def cursor(self) -> FakeSqlServerCursor:
                return self.cursor_value

            def rollback(self) -> None:
                self.rolled_back = True

            def close(self) -> None:
                self.closed = True

        sqlserver_connections: list[FakeSqlServerConnection] = []

        def sqlserver_connect(*_args: object, **_kwargs: object) -> FakeSqlServerConnection:
            connection = FakeSqlServerConnection([("SVC-1",)])
            sqlserver_connections.append(connection)
            return connection

        with patch.dict(sys.modules, {"pyodbc": SimpleNamespace(connect=sqlserver_connect)}):
            self.assertEqual(
                [{"service_id": "SVC-1"}],
                billing_query._execute_sqlserver(
                    "dsn",
                    "select :service_id",
                    {"service_id": "SVC-1"},
                    timeout_seconds=2,
                    row_limit=1,
                ),
            )
        self.assertTrue(sqlserver_connections[0].rolled_back)
        self.assertTrue(sqlserver_connections[0].closed)

        limited_connection = FakeSqlServerConnection([("one",), ("two",)])
        with patch.dict(
            sys.modules,
            {"pyodbc": SimpleNamespace(connect=lambda *_args, **_kwargs: limited_connection)},
        ):
            with self.assertRaisesRegex(RuntimeError, "row_limit_exceeded"):
                billing_query._execute_sqlserver(
                    "dsn", "select 1", {}, timeout_seconds=1, row_limit=1
                )
        self.assertTrue(limited_connection.rolled_back)
        self.assertTrue(limited_connection.closed)

        class FakePostgresCursor:
            def __init__(self, rows: list[dict]) -> None:
                self.rows = rows

            def __enter__(self) -> "FakePostgresCursor":
                return self

            def __exit__(self, *_args: object) -> None:
                return None

            def execute(self, *_args: object) -> None:
                return None

            def fetchmany(self, _count: int) -> list[dict]:
                return self.rows

        class FakePostgresConnection:
            def __init__(self, rows: list[dict]) -> None:
                self.rows = rows

            def __enter__(self) -> "FakePostgresConnection":
                return self

            def __exit__(self, *_args: object) -> None:
                return None

            def execute(self, *_args: object) -> None:
                return None

            def cursor(self) -> FakePostgresCursor:
                return FakePostgresCursor(self.rows)

        postgres_rows = [{"service_id": "SVC-1"}]
        fake_psycopg = SimpleNamespace(
            connect=lambda *_args, **_kwargs: FakePostgresConnection(postgres_rows)
        )
        with patch.dict(
            sys.modules,
            {
                "psycopg": fake_psycopg,
                "psycopg.rows": SimpleNamespace(dict_row=object()),
            },
        ):
            self.assertEqual(
                postgres_rows,
                billing_query._execute_postgres(
                    "dsn", "select 1", {}, timeout_seconds=1, row_limit=1
                ),
            )
            postgres_rows[:] = [{"service_id": "one"}, {"service_id": "two"}]
            with self.assertRaisesRegex(RuntimeError, "row_limit_exceeded"):
                billing_query._execute_postgres(
                    "dsn", "select 1", {}, timeout_seconds=1, row_limit=1
                )

        with tempfile.TemporaryDirectory() as tmp:
            database = Path(tmp) / "billing.db"
            import sqlite3

            connection = sqlite3.connect(database)
            connection.execute("CREATE TABLE candidates (service_id TEXT)")
            connection.executemany("INSERT INTO candidates VALUES (?)", [("one",), ("two",)])
            connection.commit()
            connection.close()
            with self.assertRaisesRegex(RuntimeError, "row_limit_exceeded"):
                billing_query._execute_sqlite(
                    str(database),
                    "SELECT service_id FROM candidates",
                    {},
                    timeout_seconds=1,
                    row_limit=1,
                )

        with patch.dict(
            os.environ,
            {"NEXON_RECON_BILLING_DSN": "dsn", "NEXON_RECON_BILLING_MODE": "azure_sql"},
            clear=False,
        ), patch.object(billing_query, "_execute_sqlserver", return_value=[{"ok": True}]):
            self.assertEqual([{"ok": True}], billing_query._execute_query("select 1", {}))

    def test_billing_main_initial_and_investigation_round(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            config = root / "config.yaml"
            normalized = root / "normalized.json"
            output = root / "candidates.json"
            query_log = root / "query-log.json"
            write_config(config)
            write_json(normalized, normalized_payload())
            row = {
                "provider": "AAPT",
                "provider_account": "ACC-1",
                "transaction_date": "2026-06-15",
                "service_id": "SVC-1",
                "candidate_id": "candidate-1",
            }
            args = [
                "--config",
                config,
                "--normalized",
                normalized,
                "--sql",
                VALID_SQL,
                "--output",
                output,
                "--query-log",
                query_log,
            ]
            with patch.object(billing_query, "_execute_query", return_value=[row]):
                self.assertEqual(0, call_main(billing_query, args))
            initial_log = json.loads(query_log.read_text(encoding="utf-8"))
            self.assertEqual(1, len(initial_log))

            line_ids = root / "line-ids.json"
            exception = root / "exception.json"
            audit = root / "audit.json"
            write_json(line_ids, {"line_ids": ["line-1"]})
            write_json(
                exception,
                {
                    "run_id": RUN_ID,
                    "rows": [{"line_id": "line-1"}],
                    "query_log_identity": {
                        "path": str(query_log.resolve()),
                        "chunk_count": 1,
                        "sha256": billing_query._json_payload_hash(initial_log),
                    },
                },
            )
            write_json(audit, {"run_id": RUN_ID})
            investigation_args = args + [
                "--line-ids-file",
                line_ids,
                "--exception-input",
                exception,
                "--audit-manifest",
                audit,
                "--query-round",
                "1",
                "--query-budget",
                "2",
            ]
            with patch.object(billing_query, "_execute_query", return_value=[row]):
                self.assertEqual(0, call_main(billing_query, investigation_args))
            self.assertEqual(2, len(json.loads(query_log.read_text(encoding="utf-8"))))
            self.assertEqual(
                2,
                json.loads(audit.read_text(encoding="utf-8"))["query_logs"][0][
                    "chunk_count"
                ],
            )
            write_json(audit, {"run_id": "OTHER"})
            second_round_args = [
                value if index != investigation_args.index("1") else "2"
                for index, value in enumerate(investigation_args)
            ]
            with patch.object(billing_query, "_execute_query", return_value=[row]):
                with self.assertRaisesRegex(RuntimeError, "audit manifest run_id"):
                    call_main(billing_query, second_round_args)

    def test_billing_main_scope_and_budget_rejections(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            config = root / "config.yaml"
            normalized = root / "normalized.json"
            output = root / "out.json"
            query_log = root / "query.json"
            line_ids = root / "ids.json"
            exception = root / "exception.json"
            audit = root / "audit.json"
            write_config(config)

            def invoke(extra: list[object] | None = None) -> int:
                return call_main(
                    billing_query,
                    [
                        "--config",
                        config,
                        "--normalized",
                        normalized,
                        "--sql",
                        VALID_SQL,
                        "--output",
                        output,
                        "--query-log",
                        query_log,
                        *(extra or []),
                    ],
                )

            write_json(normalized, {"lines": {}})
            with self.assertRaisesRegex(RuntimeError, "must be a list"):
                invoke()
            mixed = normalized_payload()
            mixed["lines"].append({**mixed["lines"][0], "run_id": "OTHER", "line_id": "line-2"})
            write_json(normalized, mixed)
            with self.assertRaisesRegex(RuntimeError, "one run/provider"):
                invoke()

            write_json(normalized, normalized_payload())
            for payload, message in (
                ([], "non-empty list"),
                (["line-1", "line-1"], "unique"),
                (["unknown"], "unknown rows"),
            ):
                write_json(line_ids, payload)
                with self.subTest(message=message):
                    with self.assertRaisesRegex(RuntimeError, message):
                        invoke(["--line-ids-file", line_ids])

            write_json(line_ids, {"line_ids": ["line-1"]})
            with self.assertRaisesRegex(RuntimeError, "exception-input"):
                invoke(["--line-ids-file", line_ids])

            base_exception = {
                "run_id": RUN_ID,
                "rows": [{"line_id": "line-1"}],
                "query_log_identity": {
                    "path": str(query_log.resolve()),
                    "chunk_count": 1,
                    "sha256": billing_query._json_payload_hash([{"initial": True}]),
                },
            }
            write_json(query_log, [{"initial": True}])
            write_json(audit, {"run_id": RUN_ID})

            cases = [
                ({**base_exception, "run_id": "OTHER"}, [], "run_id does not match"),
                (base_exception, [], "audit-manifest"),
                (
                    {**base_exception, "rows": [{"line_id": "other"}]},
                    ["--audit-manifest", audit],
                    "unresolved subset",
                ),
                (
                    {
                        **base_exception,
                        "query_log_identity": {
                            **base_exception["query_log_identity"],
                            "path": str(root / "other.json"),
                        },
                    },
                    ["--audit-manifest", audit],
                    "query log path",
                ),
                (
                    base_exception,
                    ["--audit-manifest", audit],
                    "require round and budget",
                ),
                (
                    base_exception,
                    [
                        "--audit-manifest",
                        audit,
                        "--query-round",
                        "1",
                        "--query-budget",
                        "1",
                    ],
                    "configured limit",
                ),
                (
                    {
                        **base_exception,
                        "query_log_identity": {
                            **base_exception["query_log_identity"],
                            "chunk_count": 0,
                        },
                    },
                    [
                        "--audit-manifest",
                        audit,
                        "--query-round",
                        "1",
                        "--query-budget",
                        "2",
                    ],
                    "provenance is invalid",
                ),
            ]
            for exception_payload, extra, message in cases:
                write_json(exception, exception_payload)
                with self.subTest(message=message):
                    with self.assertRaisesRegex(RuntimeError, message):
                        invoke(
                            [
                                "--line-ids-file",
                                line_ids,
                                "--exception-input",
                                exception,
                                *extra,
                            ]
                        )

            with self.assertRaisesRegex(RuntimeError, "only with --line-ids-file"):
                invoke(["--query-round", "1"])

            write_json(exception, base_exception)
            write_json(
                query_log,
                [{"initial": True}, {"query_round": 1}],
            )
            round_exception = copy.deepcopy(base_exception)
            round_exception["query_log_identity"]["sha256"] = billing_query._json_payload_hash(
                [{"initial": True}]
            )
            write_json(exception, round_exception)
            with self.assertRaisesRegex(RuntimeError, "exceeds the evidence budget"):
                invoke(
                    [
                        "--line-ids-file",
                        line_ids,
                        "--exception-input",
                        exception,
                        "--audit-manifest",
                        audit,
                        "--query-round",
                        "1",
                        "--query-budget",
                        "2",
                    ]
                )

            write_config(config, candidate_limit=1)
            write_json(query_log, [])
            duplicate_rows = [
                {
                    "provider": "AAPT",
                    "provider_account": "ACC-1",
                    "transaction_date": "2026-06-15",
                    "service_id": "SVC-1",
                    "candidate_id": value,
                }
                for value in ("one", "two")
            ]
            with patch.object(billing_query, "_execute_query", return_value=duplicate_rows):
                with self.assertRaisesRegex(RuntimeError, "candidate_limit_exceeded"):
                    invoke()

    def test_core_persistence_validation_and_cli_edges(self) -> None:
        normalized, candidates, matches = persistence_payloads()
        run_path = f"/run/{RUN_ID}"

        with self.assertRaisesRegex(RuntimeError, "run_id is required"):
            core_persistence._required_text("", "run_id")

        class MissingRowCursor:
            lastrowid = 0

            def execute(self, sql: str, _values: tuple) -> None:
                import sqlite3

                if sql.startswith("INSERT"):
                    raise sqlite3.IntegrityError()

            def fetchone(self) -> None:
                return None

        with self.assertRaises(Exception):
            core_persistence._insert_or_get(
                MissingRowCursor(),
                "INSERT",
                "SELECT",
                ("key",),
                {},
            )

        malformed = copy.deepcopy(normalized)
        malformed["lines"] = {}
        mutations = [
            (malformed, candidates, matches, run_path, "payloads are malformed"),
            (
                {
                    **normalized,
                    "lines": normalized["lines"]
                    + [{**normalized["lines"][0], "run_id": "OTHER", "line_id": "line-2"}],
                },
                candidates,
                matches,
                run_path,
                "one run must map",
            ),
            (normalized, candidates, matches, "/run/OTHER", "run_path does not match"),
            (
                normalized,
                {**candidates, "provider": "Other"},
                matches,
                run_path,
                "candidate evidence identity",
            ),
            (
                normalized,
                candidates,
                {"rows": [{**matches["rows"][0], "provider": "Other"}]},
                run_path,
                "match rows are not bound",
            ),
            (
                {
                    **normalized,
                    "invoice_headers": [
                        {**normalized["invoice_headers"][0], "request_key": "not-scoped"}
                    ],
                },
                candidates,
                matches,
                run_path,
                "request keys are not run-scoped",
            ),
            (
                normalized,
                candidates,
                {"rows": []},
                run_path,
                "every parser line",
            ),
            (
                {
                    **normalized,
                    "lines": [
                        {
                            **normalized["lines"][0],
                            "invoice_identity": "missing-invoice",
                        }
                    ],
                },
                candidates,
                matches,
                run_path,
                "missing invoice header",
            ),
            (
                normalized,
                {
                    **candidates,
                    "candidates_by_line": {"unknown": []},
                },
                matches,
                run_path,
                "unknown line",
            ),
        ]
        for normalized_value, candidates_value, matches_value, path, message in mutations:
            with self.subTest(message=message), tempfile.TemporaryDirectory() as tmp:
                with self.assertRaisesRegex(RuntimeError, message):
                    core_persistence.persist_shadow_run(
                        dsn=str(Path(tmp) / "shadow.db"),
                        normalized=normalized_value,
                        candidates=candidates_value,
                        matches=matches_value,
                        provider_account_id=7,
                        run_path=path,
                    )

        no_snapshot = copy.deepcopy(matches)
        no_snapshot["rows"][0]["candidate_snapshot"] = []
        no_candidates = {
            "run_id": RUN_ID,
            "provider": "AAPT",
            "candidates_by_line": {},
        }
        with tempfile.TemporaryDirectory() as tmp:
            persisted, _manifest = core_persistence.persist_shadow_run(
                dsn=str(Path(tmp) / "shadow.db"),
                normalized=normalized,
                candidates=no_candidates,
                matches=no_snapshot,
                provider_account_id=7,
                run_path=run_path,
            )
            self.assertEqual("", persisted["rows"][0]["GenericNexonBillingId"])

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            config = root / "config.yaml"
            write_config(config)
            normalized_file = root / "normalized.json"
            candidates_file = root / "candidates.json"
            matches_file = root / "matches.json"
            output = root / "persisted.json"
            manifest = root / "manifest.json"
            write_json(normalized_file, normalized)
            write_json(candidates_file, candidates)
            write_json(matches_file, matches)
            argv = [
                "--config",
                config,
                "--normalized",
                normalized_file,
                "--candidates",
                candidates_file,
                "--matches",
                matches_file,
                "--output",
                output,
                "--manifest",
                manifest,
                "--provider-account-id",
                "7",
                "--run-path",
                run_path,
            ]
            with patch.dict(os.environ, {"NEXON_RECON_CORE_MODE": "sqlite_shadow"}, clear=True):
                with self.assertRaisesRegex(RuntimeError, "DSN is missing"):
                    call_main(core_persistence, argv)
            with patch.dict(
                os.environ,
                {"NEXON_RECON_CORE_MODE": "other", "NEXON_RECON_CORE_DSN": "dsn"},
                clear=True,
            ):
                with self.assertRaisesRegex(RuntimeError, "must be sqlite_shadow"):
                    call_main(core_persistence, argv)
            with patch.dict(
                os.environ,
                {
                    "NEXON_RECON_CORE_MODE": "sqlite_shadow",
                    "NEXON_RECON_CORE_DSN": str(root / "shadow.db"),
                },
                clear=True,
            ):
                self.assertEqual(0, call_main(core_persistence, argv))

    def test_optional_update_notification_report_and_common_edges(self) -> None:
        original_import = builtins.__import__
        with patch.object(
            builtins,
            "__import__",
            side_effect=lambda name, *args, **kwargs: (
                (_ for _ in ()).throw(ImportError("openpyxl"))
                if name == "openpyxl"
                else original_import(name, *args, **kwargs)
            ),
        ):
            with self.assertRaisesRegex(RuntimeError, "openpyxl is required"):
                optional_db_update._read_rows(Path("missing.xlsx"))

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            missing_sheet = root / "missing-sheet.xlsx"
            workbook = Workbook()
            workbook.active.title = "Other"
            workbook.save(missing_sheet)
            with self.assertRaisesRegex(RuntimeError, "missing Result sheet"):
                optional_db_update._read_rows(missing_sheet)

            empty = root / "empty.xlsx"
            workbook = Workbook()
            workbook.active.title = "Result"
            workbook.save(empty)
            workbook.close()
            self.assertEqual([], optional_db_update._read_rows(empty))

            report = root / "report.xlsx"
            report.write_bytes(b"report")
            approval = root / "approval.json"
            write_json(approval, {})
            with self.assertRaisesRegex(RuntimeError, "missing fields"):
                optional_db_update._approval(approval, report)
            base_approval = {
                "run_id": RUN_ID,
                "report_id": hashlib.sha256(report.read_bytes()).hexdigest(),
                "approved_row_ids": [],
                "approved_by": "reviewer",
                "approved_at": "2026-07-09T00:00:00Z",
                "eligibility_policy_version": "v1",
                "dry_run_hash": "hash",
                "change_ticket": "CHG-1",
                "batch_idempotency_key": "batch-1",
            }
            write_json(approval, base_approval)
            with self.assertRaisesRegex(RuntimeError, "requires approved_row_ids"):
                optional_db_update._approval(approval, report)
            write_json(
                approval,
                {**base_approval, "approved_row_ids": ["line-1"], "report_id": "wrong"},
            )
            with self.assertRaisesRegex(RuntimeError, "report_id does not match"):
                optional_db_update._approval(approval, report)

            reviewed = root / "reviewed.xlsx"
            workbook = Workbook()
            result = workbook.active
            result.title = "Result"
            result.append(["line_id", "human_verified_status", "human_verified_invoice_number"])
            result.append(["line-1", "verified", "INV-1"])
            workbook.save(reviewed)
            workbook.close()
            config = root / "config.yaml"
            write_config(config, db_update_enabled=True)
            write_json(
                approval,
                {
                    **base_approval,
                    "report_id": hashlib.sha256(reviewed.read_bytes()).hexdigest(),
                    "approved_row_ids": ["absent-line"],
                },
            )
            with self.assertRaisesRegex(RuntimeError, "row IDs that are absent"):
                call_main(
                    optional_db_update,
                    [
                        "--config",
                        config,
                        "--refined-report",
                        reviewed,
                        "--audit-output",
                        root / "db-audit.json",
                        "--approval-artifact",
                        approval,
                        "--dry-run",
                    ],
                )

            audit_path = root / "run" / "manifest" / "audit_manifest.json"
            write_json(audit_path, {"run_id": RUN_ID})
            with self.assertRaisesRegex(RuntimeError, "requires message id"):
                call_main(
                    record_notification,
                    [
                        "--run-root",
                        root / "run",
                        "--status",
                        "sent",
                        "--recipient-count",
                        "1",
                    ],
                )

            row = {column: "" for column in common.RAW_WORKBOOK_COLUMNS}
            row.update(
                {
                    "ReconMatchStatus": "Matched",
                    "candidate_snapshot": [],
                    "deterministic_evidence_summary": "evidence",
                    "Adjustment": "1.25",
                    "Reason": "credit",
                }
            )
            blank_policy = {"auto_matched": "blank", "max_chars": 40}
            refined = write_reports._agent_defaults(row, blank_policy)
            self.assertEqual("", refined["agent_evidence_summary"])
            output = root / "report-output.xlsx"
            write_reports._write_workbook(
                output,
                [row, {**row, "Adjustment": "2.75", "Reason": "credit"}],
                common.RAW_WORKBOOK_COLUMNS,
                run_path="/run",
                period="2026-06",
            )
            with self.assertRaisesRegex(RuntimeError, "Invalid Adjustment"):
                write_reports._write_workbook(
                    root / "bad.xlsx",
                    [{**row, "Adjustment": "not-money"}],
                    common.RAW_WORKBOOK_COLUMNS,
                    run_path="/run",
                    period="2026-06",
                )
            manifest_path = root / "report-manifest.json"
            payload = write_reports.write_reports(
                raw_rows=[row],
                refined_input_rows=[row],
                raw_output=root / "raw.xlsx",
                refined_output=None,
                manifest=manifest_path,
                config={"reports": {"evidence_summary": blank_policy}},
                run_path="/run",
                period="2026-06",
            )
            self.assertIsNone(payload["refined_output"])
            matches = root / "matches.json"
            write_json(matches, {"rows": [row]})
            self.assertEqual(
                0,
                call_main(
                    write_reports,
                    [
                        "--config",
                        root / "missing-config.yaml",
                        "--matches",
                        matches,
                        "--raw-output",
                        root / "cli-raw.xlsx",
                        "--manifest",
                        root / "cli-manifest.json",
                    ],
                ),
            )
            real_import = builtins.__import__

            def block_openpyxl(name: str, *args: object, **kwargs: object) -> object:
                if name == "openpyxl":
                    raise ImportError(name)
                return real_import(name, *args, **kwargs)

            with patch.object(builtins, "__import__", side_effect=block_openpyxl):
                with self.assertRaisesRegex(RuntimeError, "openpyxl is required"):
                    write_reports._write_workbook(
                        root / "unavailable.xlsx",
                        [],
                        common.RAW_WORKBOOK_COLUMNS,
                        run_path="/run",
                        period="2026-06",
                    )
            gc.collect()

        with self.assertRaisesRegex(RuntimeError, "audit_required"):
            common.require_audit({"billing": {}})
        with self.assertRaisesRegex(ValueError, "must be an integer"):
            common.positive_limit({"limits": {"value": "bad"}}, "value", 1)
        self.assertEqual(1, common.positive_limit({"limits": []}, "value", 1))
        with self.assertRaisesRegex(ValueError, "greater than zero"):
            common.positive_limit({"limits": {"value": 0}}, "value", 1)
        self.assertEqual("ab", common.normalize_evidence_summary("abcd", 2))
        self.assertEqual("abcdefg...", common.normalize_evidence_summary("abcdefghijk", 10))

    def test_unpack_state_intake_and_provider_edges(self) -> None:
        class FakeInfo:
            def __init__(
                self,
                name: str,
                *,
                file_size: int,
                compress_size: int,
                external_attr: int = 0,
            ) -> None:
                self.filename = name
                self.file_size = file_size
                self.compress_size = compress_size
                self.external_attr = external_attr

            def is_dir(self) -> bool:
                return False

        class FakeArchive:
            def __init__(self, info: FakeInfo) -> None:
                self.info = info

            def __enter__(self) -> "FakeArchive":
                return self

            def __exit__(self, *_args: object) -> None:
                return None

            def infolist(self) -> list[FakeInfo]:
                return [self.info]

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            archive = root / "archive.zip"
            archive.write_bytes(b"archive")
            cases = [
                (
                    FakeInfo("large.txt", file_size=11, compress_size=10),
                    {"max_total_expanded_bytes": 10},
                    "expanded size",
                ),
                (
                    FakeInfo("zero.txt", file_size=1, compress_size=0),
                    {},
                    "zero compressed size",
                ),
                (
                    FakeInfo("ratio.txt", file_size=101, compress_size=1),
                    {"max_compression_ratio": 100},
                    "compression ratio",
                ),
                (
                    FakeInfo(
                        "link.txt",
                        file_size=1,
                        compress_size=1,
                        external_attr=(stat.S_IFLNK | 0o777) << 16,
                    ),
                    {},
                    "symlink",
                ),
            ]
            for info, limits, message in cases:
                with self.subTest(message=message), patch.object(
                    safe_unpack.zipfile, "ZipFile", return_value=FakeArchive(info)
                ):
                    inventory = safe_unpack.extract_zip(archive, root / message, **limits)
                    self.assertIn(message, inventory["blocked"][0]["reason"])

            state_path = root / "state.json"
            run_state.create_state(
                state_path,
                run_id=RUN_ID,
                provider="AAPT",
                run_mode="reconciliation",
                source_identity="source",
            )
            with self.assertRaisesRegex(ValueError, "Unknown run stage"):
                run_state.update_stage(state_path, "unknown", "running")
            with self.assertRaisesRegex(ValueError, "Unknown stage status"):
                run_state.update_stage(state_path, "provider_parsing", "unknown")
            with self.assertRaisesRegex(ValueError, "Unknown run status"):
                run_state.finalize_state(state_path, "unknown")
            self.assertEqual("completed", run_state.finalize_state(state_path, "completed")["run_status"])

            source = root / "invoice.csv"
            source.write_text("invoice", encoding="utf-8")
            config = {
                "features": {"db_update_enabled": False},
                "provider_api_adapters": {"aapt": False},
            }
            with self.assertRaisesRegex(RuntimeError, "setup_incomplete"):
                intake_run.create_run(
                    config=config,
                    provider="AAPT",
                    source_file=source,
                    result_root=root / "missing-result",
                    intake_mode="manual_upload",
                    source_identity=None,
                    copy_source=True,
                )

            args = Namespace(
                account_id="ACC",
                invoice_id="INV",
                document_id=None,
                output_dir=root,
                output_name="invoice.pdf",
                manifest=None,
            )

            def fake_download(_url: str, _token: str, destination: Path) -> None:
                destination.write_bytes(b"pdf")

            with patch.dict(
                os.environ,
                {
                    "NEXON_RECON_PROVIDER_API_CLIENT_ID_EQUINIX": "client",
                    "NEXON_RECON_PROVIDER_API_CLIENT_SECRET_EQUINIX": "secret",
                },
                clear=True,
            ), patch.object(
                provider_api_download,
                "_http_json",
                return_value={"access_token": "token"},
            ), patch.object(
                provider_api_download,
                "_http_download",
                side_effect=fake_download,
            ):
                self.assertEqual(
                    root / "invoice.pdf",
                    provider_api_download._equinix_download(args, {}),
                )


if __name__ == "__main__":
    unittest.main()

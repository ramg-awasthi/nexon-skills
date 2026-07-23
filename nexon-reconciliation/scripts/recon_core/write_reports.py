from __future__ import annotations

import argparse
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from .common import (
    APPROVED_REFINED_COLUMNS,
    DEFAULT_CONFIG_PATH,
    EXCLUDED_PHASE1_COLUMNS,
    RAW_WORKBOOK_COLUMNS,
    evidence_summary_policy,
    load_config,
    normalize_evidence_summary,
    read_json,
    write_json,
)

BASE_COLUMNS = RAW_WORKBOOK_COLUMNS


def ordered_columns(rows: list[dict[str, Any]], preferred: list[str]) -> list[str]:
    columns = list(dict.fromkeys(preferred))
    for row in rows:
        for column in row:
            if column not in columns and column not in EXCLUDED_PHASE1_COLUMNS:
                columns.append(column)
    return columns


def with_refined_defaults(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    policy = {"auto_matched": "short", "max_chars": 160}
    return [_agent_defaults(row, policy) for row in rows]


def _agent_defaults(row: dict[str, Any], policy: dict[str, Any]) -> dict[str, Any]:
    output = dict(row)
    matched = row.get("ReconMatchStatus") == "Matched" or row.get("agent_match_status") == "auto_matched"
    output.setdefault("agent_match_status", "auto_matched" if matched else "needs_review")
    output.setdefault(
        "agent_match_rule",
        row.get("deterministic_match_rule", "") if matched else "",
    )
    summary = str(row.get("deterministic_evidence_summary", "")) if matched else ""
    if matched and policy["auto_matched"] == "blank":
        summary = ""
    output.setdefault("agent_evidence_summary", normalize_evidence_summary(summary, policy["max_chars"]))
    output.setdefault("agent_review_required", not matched)
    candidate = row.get("candidate_snapshot", {})
    if not isinstance(candidate, dict):
        candidate = {}
    output.setdefault("agent_suggested_customer_account", candidate.get("customer_account", "") if matched else "")
    output.setdefault("agent_suggested_subscription_id", candidate.get("subscription_id", "") if matched else "")
    output.setdefault("agent_suggested_invoice_number", candidate.get("invoice_number", "") if matched else "")
    output.setdefault("agent_suggested_service_id", candidate.get("service_id", "") if matched else "")
    output.setdefault("human_verified_status", "not_reviewed")
    output.setdefault("human_verified_by", "")
    output.setdefault("human_verified_at", "")
    output.setdefault("human_verified_invoice_number", "")
    return output


def _write_workbook(
    output_path: Path,
    rows: list[dict[str, Any]],
    columns: list[str],
    *,
    run_path: str,
    period: str,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to generate reconciliation workbooks.") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    result = workbook.active
    result.title = "Result"
    result.append(columns)
    for row in rows:
        result.append([row.get(column, "") for column in columns])
    result.auto_filter.ref = result.dimensions
    result.freeze_panes = "A2"

    validation_columns = {
        "Dispute or Not": '"Yes,No"',
        "ManualMatch": '"Yes,No"',
        "Customer Billing Initiated": '"Yes,No"',
        "Service Cancellation Initiated": '"Yes,No"',
    }
    for name, formula in validation_columns.items():
        if name not in columns:
            continue
        column_index = columns.index(name) + 1
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        result.add_data_validation(validation)
        validation.add(f"{result.cell(2, column_index).coordinate}:{result.cell(max(2, len(rows) + 1), column_index).coordinate}")

    adjustment_total = Decimal("0")
    adjustment_reasons: list[str] = []
    for row in rows:
        value = str(row.get("Adjustment", "") or "").replace(",", "").strip()
        if value:
            try:
                adjustment_total += Decimal(value)
            except InvalidOperation as exc:
                raise RuntimeError(f"Invalid Adjustment value in report input: {value!r}") from exc
        reason = str(row.get("Reason", "") or "").strip()
        if reason and reason not in adjustment_reasons:
            adjustment_reasons.append(reason)
    adjustment = workbook.create_sheet("Adjustment")
    adjustment.append(
        [
            "Adjustment:",
            str(adjustment_total.quantize(Decimal("0.01"))),
            "Reason:",
            "; ".join(adjustment_reasons),
        ]
    )

    do_not_change = workbook.create_sheet("Do not change")
    do_not_change.append(["RunPath", run_path])
    do_not_change.append(["ReconciliationPeriod", period])

    workbook.save(output_path)


def write_reports(
    *,
    raw_rows: list[dict[str, Any]],
    refined_input_rows: list[dict[str, Any]],
    raw_output: Path,
    refined_output: Path | None,
    manifest: Path,
    config: dict[str, Any],
    run_path: str,
    period: str,
) -> dict[str, Any]:
    if len(raw_rows) != len(refined_input_rows):
        raise RuntimeError("Raw and refined report inputs must have the same row count.")
    if any(EXCLUDED_PHASE1_COLUMNS.intersection(row.keys()) for row in refined_input_rows + raw_rows):
        raise RuntimeError("Excluded runtime columns leaked into report input.")

    raw_workbook_rows = [{column: row.get(column, "") for column in RAW_WORKBOOK_COLUMNS} for row in raw_rows]
    policy = evidence_summary_policy(config)
    refined_rows = [_agent_defaults(row, policy) for row in refined_input_rows]
    refined_columns = RAW_WORKBOOK_COLUMNS + APPROVED_REFINED_COLUMNS
    refined_workbook_rows = [{column: row.get(column, "") for column in refined_columns} for row in refined_rows]

    _write_workbook(raw_output, raw_workbook_rows, RAW_WORKBOOK_COLUMNS, run_path=run_path, period=period)
    if refined_output:
        _write_workbook(refined_output, refined_workbook_rows, refined_columns, run_path=run_path, period=period)

    payload = {
        "raw_output": str(raw_output),
        "refined_output": str(refined_output) if refined_output else None,
        "row_count": len(raw_rows),
        "raw_columns": RAW_WORKBOOK_COLUMNS,
        "refined_columns": refined_columns if refined_output else [],
        "refined_added_columns": APPROVED_REFINED_COLUMNS if refined_output else [],
        "refined_preserves_all_raw_fields": bool(refined_output),
        "format": "xlsx",
    }
    write_json(manifest, payload)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="Write current raw and refined reconciliation workbooks.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--matches", type=Path, required=True)
    parser.add_argument("--raw-matches", type=Path)
    parser.add_argument("--raw-output", type=Path, required=True)
    parser.add_argument("--refined-output", type=Path)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--run-path", default="")
    parser.add_argument("--period", default="")
    args = parser.parse_args()

    refined_input_rows = read_json(args.matches).get("rows", [])
    raw_rows = read_json(args.raw_matches).get("rows", []) if args.raw_matches else refined_input_rows
    config = load_config(args.config) if args.config.is_file() else {
        "reports": {"evidence_summary": {"auto_matched": "short", "max_chars": 160}}
    }
    write_reports(
        raw_rows=raw_rows,
        refined_input_rows=refined_input_rows,
        raw_output=args.raw_output,
        refined_output=args.refined_output,
        manifest=args.manifest,
        config=config,
        run_path=args.run_path,
        period=args.period,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

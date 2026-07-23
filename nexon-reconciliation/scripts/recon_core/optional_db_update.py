from __future__ import annotations

import argparse
import hashlib
from pathlib import Path

from .common import DEFAULT_CONFIG_PATH, HUMAN_VERIFIED_STATUS_VALUES, load_config, read_json, write_json

UPDATE_STATUS_VALUES = {"verified", "deferred"}
REQUIRE_INVOICE_NUMBER_FOR_STATUSES = {"verified"}
PARTIAL_DEFERRED_UPDATE_ACTION = "write_partial_deferred_update"
HUMAN_VERIFICATION_UPDATE_ACTION = "write_human_verification_update"


def _read_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read the reviewed refined workbook.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Result" not in workbook.sheetnames:
        raise RuntimeError("Refined workbook is missing Result sheet.")
    values = list(workbook["Result"].iter_rows(values_only=True))
    if not values:
        return []
    header = [str(value or "") for value in values[0]]
    return [
        {header[index]: str(value or "") for index, value in enumerate(row)}
        for row in values[1:]
        if any(value not in (None, "") for value in row)
    ]


def _approval(path: Path, refined_report: Path) -> dict:
    approval = read_json(path)
    required = {
        "run_id",
        "report_id",
        "approved_row_ids",
        "approved_by",
        "approved_at",
        "eligibility_policy_version",
        "dry_run_hash",
        "change_ticket",
        "batch_idempotency_key",
    }
    missing = sorted(required - set(approval))
    if missing:
        raise RuntimeError(f"DB update approval artifact is missing fields: {missing}")
    if not isinstance(approval["approved_row_ids"], list) or not approval["approved_row_ids"]:
        raise RuntimeError("DB update approval artifact requires approved_row_ids.")
    digest = hashlib.sha256(refined_report.read_bytes()).hexdigest()
    if approval["report_id"] != digest:
        raise RuntimeError("DB update approval report_id does not match the reviewed workbook.")
    return approval


def _build_update_plan(rows: list[dict[str, str]], config: dict) -> dict:
    policy = config.get("db_update_policy", {})
    allow_deferred_without_invoice = policy.get("allow_deferred_without_invoice_number") is True

    planned_updates: list[dict] = []
    skipped_rows: list[dict] = []
    for index, row in enumerate(rows, start=1):
        status = row.get("human_verified_status", "")
        invoice_number = row.get("human_verified_invoice_number", "").strip()
        if status not in HUMAN_VERIFIED_STATUS_VALUES:
            raise RuntimeError(f"Invalid human_verified_status in refined report row {index}: {status}")
        if status not in UPDATE_STATUS_VALUES:
            skipped_rows.append({"row_number": index, "human_verified_status": status, "reason": "status_not_updateable"})
            continue
        if status in REQUIRE_INVOICE_NUMBER_FOR_STATUSES and not invoice_number:
            raise RuntimeError(f"Row {index} requires human_verified_invoice_number for status={status}.")

        partial_update = status == "deferred" and not invoice_number
        if partial_update and not allow_deferred_without_invoice:
            raise RuntimeError("Deferred rows without invoice number are not allowed by db_update_policy.")

        planned_updates.append(
            {
                "row_number": index,
                "line_id": row.get("line_id"),
                "provider": row.get("provider"),
                "service_id": row.get("agent_suggested_service_id") or row.get("service_id_normalized") or row.get("service_id_raw"),
                "subscription_id": row.get("agent_suggested_subscription_id"),
                "customer_account": row.get("agent_suggested_customer_account"),
                "invoice_number": invoice_number,
                "human_verified_status": status,
                "partial_update": partial_update,
                "update_action": PARTIAL_DEFERRED_UPDATE_ACTION if partial_update else HUMAN_VERIFICATION_UPDATE_ACTION,
            }
        )

    return {
        "mode": "dry_run",
        "db_update_enabled": True,
        "live_write_performed": False,
        "planned_update_count": len(planned_updates),
        "skipped_row_count": len(skipped_rows),
        "planned_updates": planned_updates,
        "skipped_rows": skipped_rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Disabled/config-gated DB update adapter.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--refined-report", type=Path, required=True)
    parser.add_argument("--audit-output", type=Path, required=True)
    parser.add_argument("--approval-artifact", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    config = load_config(args.config)
    enabled = config.get("features", {}).get("db_update_enabled") is True
    if not enabled:
        raise RuntimeError(
            "DB update is disabled for accepted resolutions. Core reconciliation "
            "persistence is a separate required runtime capability."
        )
    if args.approval_artifact is None:
        raise RuntimeError("DB update requires a controlled approval artifact.")
    if not args.dry_run:
        raise RuntimeError("DB update implementation must support and pass dry-run before live writes.")

    approval = _approval(args.approval_artifact, args.refined_report)
    rows = _read_rows(args.refined_report)
    approved_ids = {str(value) for value in approval["approved_row_ids"]}
    selected = [row for row in rows if str(row.get("line_id") or "") in approved_ids]
    selected_ids = {str(row.get("line_id") or "") for row in selected}
    if selected_ids != approved_ids:
        raise RuntimeError("DB update approval contains row IDs that are absent from the reviewed workbook.")
    plan = _build_update_plan(selected, config)
    plan["approval"] = approval
    write_json(args.audit_output, plan)
    print(args.audit_output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

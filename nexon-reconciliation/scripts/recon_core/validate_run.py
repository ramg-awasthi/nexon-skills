from __future__ import annotations

import argparse
import hashlib
import re
import zipfile
from pathlib import Path
from typing import Any

from .common import (
    AGENT_MATCH_STATUS_VALUES,
    APPROVED_REFINED_COLUMNS,
    DEFAULT_CONFIG_PATH,
    EXCLUDED_PHASE1_COLUMNS,
    HUMAN_VERIFIED_STATUS_VALUES,
    RAW_WORKBOOK_COLUMNS,
    RUN_SUBDIRS,
    evidence_summary_policy,
    load_config,
    logical_sharepoint_run_path,
    read_json,
    require_audit,
    validate_run_id,
    write_json,
)

SECRET_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in (
        r"code=",
        r"sig=",
        r"client_secret",
        r"accountkey=",
        r"sharedaccesssignature",
        r"password\s*[:=]",
    )
]


def _load_workbook(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to validate reconciliation workbooks.") from exc
    return load_workbook(path, read_only=False, data_only=False)


def workbook_rows(path: Path) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    workbook = _load_workbook(path)
    sheet_names = workbook.sheetnames
    if "Result" not in sheet_names:
        raise RuntimeError(f"Workbook is missing Result sheet: {path}")
    result = workbook["Result"]
    values = list(result.iter_rows(values_only=True))
    if not values:
        raise RuntimeError(f"Result sheet is empty: {path}")
    header = [str(value or "") for value in values[0]]
    rows = [
        {header[index]: value for index, value in enumerate(row) if index < len(header)}
        for row in values[1:]
        if any(value not in (None, "") for value in row)
    ]
    return header, rows, sheet_names


def assert_workbook_metadata(
    path: Path,
    run_root: Path,
    period: str,
    *,
    provider: str | None = None,
) -> None:
    workbook = _load_workbook(path)
    values = list(workbook["Do not change"].iter_rows(values_only=True))
    metadata = {
        str(row[0] or ""): str(row[1] or "")
        for row in values
        if len(row) >= 2 and row[0]
    }
    expected_run_path = (
        logical_sharepoint_run_path(provider, run_root) if provider else str(run_root)
    )
    if metadata.get("RunPath") != expected_run_path:
        raise RuntimeError("Workbook RunPath metadata does not match the current run.")
    if metadata.get("ReconciliationPeriod") != period:
        raise RuntimeError("Workbook reconciliation period does not match the run manifest.")


def _comparable_rows(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [
        {
            column: str(row.get(column) if row.get(column) is not None else "")
            for column in RAW_WORKBOOK_COLUMNS
        }
        for row in rows
    ]


def _artifact_text(path: Path) -> str:
    if path.suffix.lower() == ".xlsx":
        with zipfile.ZipFile(path) as archive:
            return "\n".join(
                archive.read(name).decode("utf-8", errors="ignore")
                for name in archive.namelist()
                if name.endswith(".xml")
            )
    return path.read_text(encoding="utf-8", errors="ignore")


def assert_no_secret_markers(run_root: Path) -> None:
    scan_roots = [
        run_root / "manifest",
        run_root / "logs",
        run_root / "raw-recon-report",
        run_root / "refined-recon-report",
    ]
    for root in scan_roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            try:
                text = _artifact_text(path)
            except (OSError, zipfile.BadZipFile) as exc:
                raise RuntimeError(f"Unable to scan output for secrets: {path}") from exc
            for pattern in SECRET_PATTERNS:
                if pattern.search(text):
                    raise RuntimeError(f"Secret-like marker found in output artifact: {path}")


def assert_result_path_shape(run_root: Path, run_id: str) -> None:
    provider, ymd, _time, _hash5 = run_id.split("_")
    expected_year = ymd[:4]
    expected_month = ymd[4:6]
    actual_month = run_root.parent.name
    actual_year = run_root.parent.parent.name
    actual_provider = run_root.parent.parent.parent.name
    if actual_provider != provider or actual_year != expected_year or actual_month != expected_month:
        raise RuntimeError(
            "Run path must end with <provider>/<year>/<month>/<run_id>: "
            f"expected {provider}/{expected_year}/{expected_month}/{run_id}, got {run_root}"
        )


def assert_parser_contract(
    run_root: Path, run_id: str, provider: str
) -> tuple[int, list[dict[str, Any]]]:
    warnings_path = run_root / "logs" / "parser_warnings.json"
    manifest_path = run_root / "manifest" / "parser_manifest.json"
    normalized_path = run_root / "normalized" / "provider_lines.json"
    for path in (warnings_path, manifest_path, normalized_path):
        if not path.is_file():
            raise RuntimeError(f"Missing parser artifact: {path}")
    warnings = read_json(warnings_path)
    if not isinstance(warnings, list):
        raise RuntimeError("Parser warnings artifact must be a list.")
    if any(isinstance(item, dict) and item.get("severity") == "error" for item in warnings):
        raise RuntimeError("Parser error warnings block run validation.")
    manifest = read_json(manifest_path)
    normalized = read_json(normalized_path)
    if manifest.get("run_id") != run_id or manifest.get("provider") != provider:
        raise RuntimeError("Parser manifest identity does not match the current run.")
    lines = normalized.get("lines", [])
    headers = normalized.get("invoice_headers", [])
    if not isinstance(lines, list) or not isinstance(headers, list):
        raise RuntimeError("Normalized parser output must contain invoice_headers and lines.")
    if manifest.get("accounting_complete") is not True:
        raise RuntimeError("Parser manifest does not prove complete member/row accounting.")
    source_rows = int(manifest.get("source_rows", -1))
    exclusions = int(manifest.get("documented_exclusions", -1))
    if source_rows != len(lines) + exclusions:
        raise RuntimeError("Parser source row accounting does not reconcile.")
    if int(manifest.get("parsed_rows", -1)) != len(lines):
        raise RuntimeError("Parser manifest parsed_rows does not match normalized rows.")
    if any(not row.get("line_id") or not row.get("invoice_identity") for row in lines):
        raise RuntimeError("Every normalized line requires line_id and invoice_identity.")
    if any(row.get("run_id") != run_id or row.get("provider") != provider for row in lines):
        raise RuntimeError("Normalized parser rows are not bound to the current run/provider.")
    return len(lines), warnings


def assert_parser_warnings_resolved(run_root: Path, _refined_rows: list[dict[str, Any]] | None = None) -> None:
    warnings_path = run_root / "logs" / "parser_warnings.json"
    if not warnings_path.is_file():
        raise RuntimeError(f"Missing parser warnings artifact: {warnings_path}")
    warnings = read_json(warnings_path)
    if not isinstance(warnings, list):
        raise RuntimeError("Parser warnings artifact must be a list.")
    if any(isinstance(item, dict) and item.get("severity") == "error" for item in warnings):
        raise RuntimeError("Parser error warnings block run validation.")


def assert_audit(run_root: Path, run_id: str) -> dict[str, Any]:
    audit_path = run_root / "manifest" / "audit_manifest.json"
    state_path = run_root / "manifest" / "run_state.json"
    if not audit_path.is_file():
        raise RuntimeError("Mandatory audit manifest is missing.")
    if not state_path.is_file():
        raise RuntimeError("Durable run state is missing.")
    audit = read_json(audit_path)
    state = read_json(state_path)
    if audit.get("run_id") != run_id or state.get("run_id") != run_id:
        raise RuntimeError("Audit/state run identity does not match run folder.")
    if audit.get("accepted_resolution_update_attempted") is not False:
        raise RuntimeError("Audit indicates accepted-resolution update was attempted.")
    if not isinstance(state.get("stages"), dict):
        raise RuntimeError("Run state has no stage ledger.")
    return state


def assert_reconciliation_runtime(run_root: Path, state: dict[str, Any], parsed_rows: int) -> None:
    required_files = (
        run_root / "evidence" / "billing_candidates.json",
        run_root / "logs" / "billing_query_log.json",
        run_root / "normalized" / "match_results.json",
        run_root / "manifest" / "persistence_manifest.json",
        run_root / "normalized" / "persisted_match_results.json",
    )
    missing = [str(path) for path in required_files if not path.is_file()]
    if missing:
        raise RuntimeError(f"Reconciliation runtime artifacts are missing: {missing}")
    persistence = read_json(run_root / "manifest" / "persistence_manifest.json")
    run_id = run_root.name
    provider = run_root.parent.parent.parent.name
    if persistence.get("run_id") != run_id or persistence.get("provider") != provider:
        raise RuntimeError("Persistence manifest identity does not match the current run.")
    if persistence.get("transaction") != "committed":
        raise RuntimeError("Persistence manifest does not prove a committed transaction.")
    if int(persistence.get("supplier_line_count", -1)) != parsed_rows:
        raise RuntimeError("Persistence supplier-line count does not match parser output.")
    result_count = int(persistence.get("result_count", -1))
    if result_count < parsed_rows:
        raise RuntimeError("Persistence has fewer results than parser lines.")
    persisted_rows = read_json(
        run_root / "normalized" / "persisted_match_results.json"
    ).get("rows", [])
    match_rows = read_json(run_root / "normalized" / "match_results.json").get(
        "rows", []
    )
    if len(persisted_rows) != result_count:
        raise RuntimeError("Persistence result count does not match persisted rows.")
    if {
        str(row.get("line_id", "")) for row in persisted_rows
    } != {
        str(row.get("line_id", "")) for row in match_rows
    }:
        raise RuntimeError("Persisted and matched result line identities differ.")
    query_log = read_json(run_root / "logs" / "billing_query_log.json")
    if not isinstance(query_log, list) or not query_log:
        raise RuntimeError("Billing query log must contain at least one audited query chunk.")
    candidates = read_json(run_root / "evidence" / "billing_candidates.json")
    if candidates.get("run_id") != run_id or candidates.get("provider") != provider:
        raise RuntimeError("Billing candidate evidence identity does not match the current run.")
    audit = read_json(run_root / "manifest" / "audit_manifest.json")
    audit_logs = audit.get("query_logs", [])
    if not isinstance(audit_logs, list) or not audit_logs:
        raise RuntimeError("Audit manifest is missing billing query-log provenance.")
    logged_hashes = {str(item.get("sha256")) for item in audit_logs if isinstance(item, dict)}
    query_hash = hashlib.sha256(
        (run_root / "logs" / "billing_query_log.json").read_bytes()
    ).hexdigest()
    if query_hash not in logged_hashes:
        raise RuntimeError("Billing query log hash is not bound into the audit manifest.")
    expected_ids = {
        str(row.get("line_id"))
        for row in read_json(run_root / "normalized" / "provider_lines.json").get("lines", [])
    }
    for artifact_name in ("match_results.json", "persisted_match_results.json"):
        artifact_ids = {
            str(row.get("line_id"))
            for row in read_json(run_root / "normalized" / artifact_name).get("rows", [])
        }
        if artifact_ids != expected_ids:
            raise RuntimeError(f"{artifact_name} line identities do not match parser output.")

    stages = state["stages"]
    required_completed = {
        "source_staging",
        "run_creation",
        "archive_validation",
        "provider_parsing",
        "billing_preparation",
        "deterministic_comparison",
        "supplier_persistence",
        "result_persistence",
        "raw_workbook",
    }
    incomplete = sorted(
        stage for stage in required_completed if stages.get(stage, {}).get("status") != "completed"
    )
    if incomplete:
        raise RuntimeError(f"Required reconciliation stages are not complete: {incomplete}")
    for stage in ("exception_investigation", "refined_workbook", "publication", "notification"):
        if stages.get(stage, {}).get("status") not in {"completed", "skipped"}:
            raise RuntimeError(f"Reconciliation stage is not terminal: {stage}")
    publication_status = stages.get("publication", {}).get("status")
    if publication_status == "completed":
        receipt = run_root / "manifest" / "publication_receipt.json"
        if not receipt.is_file():
            raise RuntimeError("Completed publication has no receipt.")


def assert_evidence_summary(row: dict[str, Any], *, auto_matched_mode: str, max_chars: int) -> None:
    status = str(row.get("agent_match_status", ""))
    summary = str(row.get("agent_evidence_summary", "") or "").strip()
    required = status != "excluded" and not (status == "auto_matched" and auto_matched_mode == "blank")
    if required and not summary:
        raise RuntimeError("agent_evidence_summary is required for refined report rows.")
    if "\n" in summary or "\r" in summary:
        raise RuntimeError("agent_evidence_summary must be a single line.")
    if len(summary) > max_chars:
        raise RuntimeError(f"agent_evidence_summary exceeds max_chars={max_chars}.")


def validate_workbooks(run_root: Path, config: dict[str, Any], report_manifest: dict[str, Any]) -> int:
    raw_path = Path(report_manifest.get("raw_output", "")).resolve()
    expected_raw = (run_root / "raw-recon-report" / "raw-reconciliation.xlsx").resolve()
    if raw_path != expected_raw:
        raise RuntimeError("Report manifest raw path is not the current run raw workbook.")
    if not raw_path.is_file():
        raise RuntimeError("Raw workbook is missing.")
    raw_header, raw_rows, raw_sheets = workbook_rows(raw_path)
    run_manifest = read_json(run_root / "manifest" / "run_manifest.json")
    period = str(run_manifest.get("billing_period", ""))
    provider = run_root.parent.parent.parent.name
    assert_workbook_metadata(raw_path, run_root, period, provider=provider)
    if raw_header != RAW_WORKBOOK_COLUMNS:
        raise RuntimeError("Raw Result columns do not match the exact current 35-column contract.")
    if raw_sheets != ["Result", "Adjustment", "Do not change"]:
        raise RuntimeError(f"Raw workbook sheet contract mismatch: {raw_sheets}")
    if any(column.startswith(("agent_", "human_")) for column in raw_header):
        raise RuntimeError("Agent/human columns leaked into raw workbook.")
    persisted_rows = read_json(
        run_root / "normalized" / "persisted_match_results.json"
    ).get("rows", [])
    if _comparable_rows(raw_rows) != _comparable_rows(persisted_rows):
        raise RuntimeError("Raw workbook contents do not match persisted reconciliation results.")

    refined_value = report_manifest.get("refined_output")
    if refined_value:
        refined_path = Path(refined_value).resolve()
        expected_refined = (
            run_root / "refined-recon-report" / "refined-reconciliation.xlsx"
        ).resolve()
        if refined_path != expected_refined:
            raise RuntimeError("Report manifest refined path is not the current run refined workbook.")
        if not refined_path.is_file():
            raise RuntimeError("Refined workbook is missing.")
        refined_header, refined_rows, refined_sheets = workbook_rows(refined_path)
        assert_workbook_metadata(refined_path, run_root, period, provider=provider)
        expected = RAW_WORKBOOK_COLUMNS + APPROVED_REFINED_COLUMNS
        if refined_header != expected:
            raise RuntimeError("Refined Result columns do not preserve raw fields plus approved additions.")
        if refined_sheets != ["Result", "Adjustment", "Do not change"]:
            raise RuntimeError(f"Refined workbook sheet contract mismatch: {refined_sheets}")
        if len(raw_rows) != len(refined_rows):
            raise RuntimeError("Raw/refined workbook row counts differ.")
        if _comparable_rows(refined_rows) != _comparable_rows(persisted_rows):
            raise RuntimeError("Refined workbook raw fields do not match persisted results.")
        policy = evidence_summary_policy(config)
        for row in refined_rows:
            status = str(row.get("agent_match_status", ""))
            human_status = str(row.get("human_verified_status", ""))
            if status not in AGENT_MATCH_STATUS_VALUES:
                raise RuntimeError(f"Invalid agent_match_status: {status}")
            if human_status not in HUMAN_VERIFIED_STATUS_VALUES:
                raise RuntimeError(f"Invalid human_verified_status: {human_status}")
            if human_status == "verified" and not str(row.get("human_verified_invoice_number", "") or "").strip():
                raise RuntimeError("human_verified_invoice_number is required for verified rows.")
            if any(column in refined_header for column in EXCLUDED_PHASE1_COLUMNS):
                raise RuntimeError("Excluded Phase 1 columns are present.")
            assert_evidence_summary(
                row,
                auto_matched_mode=policy["auto_matched"],
                max_chars=policy["max_chars"],
            )
    if int(report_manifest.get("row_count", -1)) != len(raw_rows):
        raise RuntimeError("Report manifest row_count does not match raw workbook.")
    return len(raw_rows)


def validate_run(run_root: Path, config: dict[str, Any], run_mode: str) -> dict[str, Any]:
    require_audit(config)
    run_id = run_root.name
    if not validate_run_id(run_id):
        raise RuntimeError(f"Invalid run_id: {run_id}")
    assert_result_path_shape(run_root, run_id)
    for subdir in RUN_SUBDIRS:
        if not (run_root / subdir).exists():
            raise RuntimeError(f"Missing run subdir: {run_root / subdir}")

    run_manifest_path = run_root / "manifest" / "run_manifest.json"
    if not run_manifest_path.is_file():
        raise RuntimeError("Missing run manifest.")
    run_manifest = read_json(run_manifest_path)
    if run_manifest.get("run_id") != run_id:
        raise RuntimeError("Run manifest run_id does not match run folder.")
    if run_manifest.get("db_update_enabled") is not False:
        raise RuntimeError("Accepted-resolution update was not proven disabled.")

    provider = run_root.parent.parent.parent.name
    parsed_rows, warnings = assert_parser_contract(run_root, run_id, provider)
    state = assert_audit(run_root, run_id)
    reported_rows = None
    if run_mode == "reconciliation":
        assert_reconciliation_runtime(run_root, state, parsed_rows)
        report_manifest_path = run_root / "manifest" / "report_manifest.json"
        if not report_manifest_path.is_file():
            raise RuntimeError("Missing report manifest.")
        reported_rows = validate_workbooks(run_root, config, read_json(report_manifest_path))
        if reported_rows != parsed_rows:
            raise RuntimeError("Parsed and reported row counts differ.")
    assert_no_secret_markers(run_root)
    return {
        "run_id": run_id,
        "run_mode": run_mode,
        "parsed_rows": parsed_rows,
        "reported_rows": reported_rows,
        "warning_count": len(warnings),
        "run_status": state.get("run_status"),
        "validation": "passed",
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate a reconciliation or parser-validation run package.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--run-root", type=Path, required=True)
    parser.add_argument("--run-mode", choices=["reconciliation", "parser_validation"], default="reconciliation")
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    result = validate_run(args.run_root, load_config(args.config), args.run_mode)
    if args.output:
        write_json(args.output, result)
    print("Run validation passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
